"""
Pure data-reading helpers for subsidiary .atbr.xlsx files — no Qt dependency,
so both the UI (consolidation_window.py) and the Excel exporter
(exporter/review_package.py) can share one implementation instead of two
copies drifting apart.

Column progression from subsidiaries:
  UNADJ  = pbc_balance
  +AJE   = adj_balance
  +RJE   = final_balance   ← what we aggregate here (book-final)
  +FTJE  = ftax_balance    ← also carried for Tax Entries column in summary
"""
from __future__ import annotations

import json
from pathlib import Path

from atbworkup.data.tax_line_categories import classify_section

# {section_name: {line_name: {"final": float, "ftax": float, "raw": float,
#                              "sort": int, "category": str}}}
SectionData = dict[str, dict[str, dict]]
# {section_name: {line_name: [{"name": str, "number": str, "final": float,
#                               "account_id": str}]}}
AccountDetail = dict[str, dict[str, list[dict]]]


def _load_data(xlsx_path: Path) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "__data" not in wb.sheetnames:
        raise ValueError("Not a valid .atbr.xlsx package")
    ws = wb["__data"]
    raw = "".join(str(cell) for row in ws.iter_rows(values_only=True)
                  for cell in row if cell is not None)
    return json.loads(raw)


def read_member_info(xlsx_path: Path) -> dict:
    """Read job metadata from a subsidiary .atbr.xlsx."""
    data = _load_data(xlsx_path)
    job = data.get("job") or {}
    return {
        "client_name": job.get("client_name", ""),
        "tax_year":    job.get("tax_year", 0),
        "entity_type": job.get("entity_type", ""),
        "status":      job.get("status", ""),
    }


def read_member_financials(
    xlsx_path: Path,
) -> tuple[SectionData, SectionData, AccountDetail, AccountDetail]:
    """
    Read accounts + entries from subsidiary .atbr.xlsx.

    Splits entries by type so we report FINAL balance (pbc + AJE + RJE):
        adj   = pbc + AJE
        final = adj + RJE   ← aggregated here

    Returns (bs_data, pl_data, bs_detail, pl_detail).
    """
    data = _load_data(xlsx_path)

    entry_status = {e["aje_id"]: e.get("status", "Open") for e in data.get("entries", [])}
    entry_type   = {e["aje_id"]: e.get("entry_type", "AJE") for e in data.get("entries", [])}

    aje_map:  dict[str, float] = {}
    rje_map:  dict[str, float] = {}
    ftje_map: dict[str, float] = {}

    for ln in data.get("entry_lines", []):
        eid = ln["aje_id"]
        if entry_status.get(eid, "Open") == "Void":
            continue
        aid  = ln["account_id"]
        amt  = float(ln["amount"])
        etype = entry_type.get(eid, "AJE")
        if etype == "AJE":
            aje_map[aid]  = aje_map.get(aid, 0.0)  + amt
        elif etype == "RJE":
            rje_map[aid]  = rje_map.get(aid, 0.0)  + amt
        elif etype == "FTJE":
            ftje_map[aid] = ftje_map.get(aid, 0.0) + amt

    tl_map = {tl["tax_line_id"]: tl for tl in data.get("tax_lines", [])}
    acct_to_tl: dict[str, str] = {
        m["account_id"]: m.get("tax_line_id") or ""
        for m in data.get("mappings", [])
    }

    bs_result: SectionData = {}
    pl_result: SectionData = {}
    bs_detail: AccountDetail = {}
    pl_detail: AccountDetail = {}

    for acct in data.get("accounts", []):
        aid  = acct["account_id"]
        pbc  = float(acct.get("pbc_balance") or 0.0)
        adj  = pbc + aje_map.get(aid, 0.0)
        final= adj + rje_map.get(aid, 0.0)
        ftax = final + ftje_map.get(aid, 0.0)
        nb   = acct.get("normal_balance", "Debit")

        tl_id = acct_to_tl.get(aid, "")
        if not tl_id or tl_id not in tl_map:
            continue
        tl      = tl_map[tl_id]
        stmt    = tl.get("financial_statement", "")

        # Balance Sheet: keep raw DR/CR sign convention (DR=+, CR=−).
        # Grand total of all BS accounts = 0 when balanced; nonzero = out of balance.
        # P&L: flip credit-normal accounts to positive so revenue displays as positive.
        def _d(v: float, _nb: str = nb) -> float:
            return -v if _nb == "Credit" else v

        if stmt == "BalanceSheet":
            disp_final = final
            disp_ftax  = ftax
        else:
            disp_final = _d(final)
            disp_ftax  = _d(ftax)
        section = tl.get("section", "")
        lname   = tl.get("line_name", "")
        sort    = tl.get("section_sort_order", 0) * 1000 + tl.get("sort_order", 0)
        # category is the stored, authoritative bucket (Revenue/COGS/OpEx/
        # ScheduleK/Asset/Liability/Equity) — falls back to the name-based
        # classifier only for tax lines saved before this column existed.
        category = tl.get("category") or classify_section(stmt, section)

        target = (bs_result if stmt == "BalanceSheet"
                  else pl_result if stmt == "ProfitAndLoss"
                  else None)
        det_target = (bs_detail if stmt == "BalanceSheet"
                      else pl_detail if stmt == "ProfitAndLoss"
                      else None)
        if target is None:
            continue
        if section not in target:
            target[section] = {}
        if lname not in target[section]:
            target[section][lname] = {
                "final": 0.0, "ftax": 0.0, "raw": 0.0, "sort": sort, "category": category,
            }
        target[section][lname]["final"] += disp_final
        target[section][lname]["ftax"]  += disp_ftax
        # Raw (never flipped) value — needed to correctly combine P&L totals
        # across sections. Revenue and expense sections are BOTH shown as
        # positive numbers for display, so blindly summing "final" across all
        # P&L sections adds expenses instead of subtracting them. The raw
        # DR+/CR- value self-cancels correctly: net income = -sum(raw).
        target[section][lname]["raw"] += final

        # Per-account detail for "Show Account Detail" toggle
        if det_target is not None:
            if section not in det_target:
                det_target[section] = {}
            if lname not in det_target[section]:
                det_target[section][lname] = []
            det_target[section][lname].append({
                "name":       acct.get("account_name", ""),
                "number":     acct.get("account_number", "") or "",
                "final":      disp_final,
                "account_id": aid,
            })

    return bs_result, pl_result, bs_detail, pl_detail


def read_member_account_index(xlsx_path: Path) -> dict[str, dict]:
    """
    Lightweight read of a subsidiary's accounts + tax-line mapping — used to
    populate the account-level elimination/CTE editor's account picker and to
    fold those lines into the combined section totals.

    Returns {account_id: {"number","name","section","stmt","normal_balance",
                           "line_code","category"}}
    """
    data = _load_data(xlsx_path)

    tl_map = {tl["tax_line_id"]: tl for tl in data.get("tax_lines", [])}
    acct_to_tl: dict[str, str] = {
        m["account_id"]: m.get("tax_line_id") or ""
        for m in data.get("mappings", [])
    }

    index: dict[str, dict] = {}
    for acct in data.get("accounts", []):
        aid = acct["account_id"]
        tl_id = acct_to_tl.get(aid, "")
        tl = tl_map.get(tl_id)
        section = tl.get("section", "") if tl else ""
        stmt    = tl.get("financial_statement", "") if tl else ""
        category = (tl.get("category") if tl else "") or classify_section(stmt, section)
        index[aid] = {
            "number":         acct.get("account_number") or "",
            "name":           acct["account_name"],
            "section":        section,
            "stmt":           stmt,
            "normal_balance": acct.get("normal_balance", "Debit"),
            "line_code":      tl.get("line_code", "") if tl else "",
            "category":       category,
        }
    return index


def merge_into(dest: SectionData, src: SectionData) -> None:
    """Merge src aggregation dict into dest in-place."""
    for section, lines in src.items():
        if section not in dest:
            dest[section] = {}
        for lname, vals in lines.items():
            if lname not in dest[section]:
                dest[section][lname] = {
                    "final": 0.0, "ftax": 0.0, "raw": 0.0, "sort": vals["sort"],
                    "category": vals.get("category", ""),
                }
            dest[section][lname]["final"] += vals["final"]
            dest[section][lname]["ftax"]  += vals["ftax"]
            dest[section][lname]["raw"]   += vals.get("raw", vals["final"])
