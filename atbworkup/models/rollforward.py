"""
Rollforward — create a new tax year job from a prior year .atbr.xlsx package.

What carries forward:
  - Accounts (same number/name/type) with PBC = prior year final adjusted balance
    for Balance Sheet accounts; PBC = 0 for P&L accounts (income resets each year).
  - Tax line mappings (account → tax line assignments).
  - Account groups and group membership (new UUIDs, same names/hierarchy).

What does NOT carry forward:
  - Journal entries (those are prior year work papers).
  - Notes.
  - Package history.

The preparer should import the new year's trial balance after rollforward, which
will update the PBC column with the client's actual books. Until then, the rolled
adjusted balances serve as a reference.
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from atbworkup.db.connection import db_connection
from atbworkup.models.job import create_workup, get_job
from atbworkup.models.activity import log_activity
from atbworkup.utils.ids import new_uuid


# ── Public entry point ────────────────────────────────────────────────────────

def create_rollforward(
    prior_xlsx_path: str | Path,
    new_atbw_path:   str | Path,
    new_metadata:    dict,
    performed_by:    str,
) -> dict:
    """
    Create a new .atbw binder rolled forward from *prior_xlsx_path*.

    Returns the new job dict (same shape as get_job()).
    Raises ValueError if the package cannot be parsed.
    """
    prior_data     = _read_package_data(Path(prior_xlsx_path))
    balances       = _compute_balances(prior_data)
    prior_job_id   = prior_data.get("job_id", "")
    new_job_id     = new_uuid()

    new_metadata = dict(new_metadata)
    new_metadata["is_rollforward"]    = 1
    new_metadata["prior_year_job_id"] = prior_job_id

    create_workup(Path(new_atbw_path), new_metadata, job_id=new_job_id)

    with db_connection(Path(new_atbw_path)) as conn:
        _populate_rollforward(
            conn, new_job_id, prior_job_id,
            prior_data, balances, performed_by,
        )

    return get_job(Path(new_atbw_path))


# ── Package reader ────────────────────────────────────────────────────────────

def _read_package_data(xlsx_path: Path) -> dict:
    """
    Extract and parse the JSON __data snapshot from a .atbr.xlsx file.
    Raises ValueError if the sheet is missing or the JSON is malformed.
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot open package: {exc}") from exc

    if "__data" not in wb.sheetnames:
        raise ValueError("This file does not contain a __data sheet. "
                         "Please select a valid .atbr.xlsx review package.")

    ws = wb["__data"]
    chunks: list[str] = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                chunks.append(str(cell))
    raw = "".join(chunks)

    try:
        return json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Package data is corrupt: {exc}") from exc


def read_prior_year_summary(xlsx_path: str | Path) -> dict:
    """
    Lightweight read — returns just the job metadata from a prior year package.
    Used by the dialog to show a summary before the user commits.
    Raises ValueError on bad file.
    """
    data = _read_package_data(Path(xlsx_path))
    job = data.get("job") or {}
    return {
        "job_id":       data.get("job_id", ""),
        "client_name":  job.get("client_name", ""),
        "entity_name":  job.get("entity_name", ""),
        "tax_year":     job.get("tax_year", 0),
        "entity_type":  job.get("entity_type", ""),
        "accounting_system": job.get("accounting_system", ""),
        "account_count": len(data.get("accounts", [])),
        "mapping_count": len(data.get("mappings", [])),
    }


# ── Balance computation ───────────────────────────────────────────────────────

def _compute_balances(data: dict) -> dict[str, dict]:
    """
    Return {account_id: {"adj": float, "final": float, "ftax": float}}
    from the __data snapshot, splitting by entry type.

    adj   = pbc + AJE
    final = adj + RJE  (book closing balance — used as rollforward PBC)
    ftax  = final + FTJE
    """
    entry_meta: dict[str, tuple[str, str]] = {}  # aje_id → (status, entry_type)
    for e in data.get("entries", []):
        entry_meta[e["aje_id"]] = (
            e.get("status", "Open"),
            e.get("entry_type", "AJE"),
        )

    by_type: dict[str, dict[str, float]] = {}
    for ln in data.get("entry_lines", []):
        status, etype = entry_meta.get(ln["aje_id"], ("Open", "AJE"))
        if status == "Void":
            continue
        aid = ln["account_id"]
        if aid not in by_type:
            by_type[aid] = {}
        by_type[aid][etype] = by_type[aid].get(etype, 0.0) + float(ln["amount"])

    result: dict[str, dict] = {}
    for acct in data.get("accounts", []):
        aid = acct["account_id"]
        pbc  = float(acct.get("pbc_balance") or 0.0)
        sums = by_type.get(aid, {})
        aje  = sums.get("AJE", 0.0)
        rje  = sums.get("RJE", 0.0)
        ftje = sums.get("FTJE", 0.0)
        adj   = round(pbc + aje,        2)
        final = round(adj + rje,        2)
        ftax  = round(final + ftje,     2)
        result[aid] = {"adj": adj, "final": final, "ftax": ftax}

    return result


# ── Populator ────────────────────────────────────────────────────────────────

_BS_TYPES = {"Asset", "Liability", "Equity"}
_PL_TYPES = {"Revenue", "Expense"}


def _populate_rollforward(
    conn,
    new_job_id:   str,
    prior_job_id: str,
    data:         dict,
    balances:     dict[str, dict],
    performed_by: str,
) -> None:
    import datetime
    now = datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    # ── Accounts ──────────────────────────────────────────────────────────
    old_to_new: dict[str, str] = {}

    for acct in data.get("accounts", []):
        old_id    = acct["account_id"]
        new_id    = new_uuid()
        old_to_new[old_id] = new_id

        acct_type = acct.get("account_type", "Asset")
        bal       = balances.get(old_id, {"adj": 0.0, "final": 0.0, "ftax": 0.0})

        # Balance sheet accounts carry their book closing balance (final = pbc+AJE+RJE).
        # P&L accounts reset to zero each year.
        if acct_type in _BS_TYPES:
            new_pbc = bal["final"]
        else:
            new_pbc = 0.0

        conn.execute(
            """INSERT OR IGNORE INTO accounts
                 (account_id, job_id, account_number, account_name, account_type,
                  pbc_balance, normal_balance, source_row, is_mapped, flag,
                  sort_order, created_at, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                new_id, new_job_id,
                acct.get("account_number"),
                acct["account_name"],
                acct_type,
                new_pbc,
                acct.get("normal_balance", "Debit"),
                acct.get("source_row"),
                acct.get("is_mapped", 0),
                None,   # clear flags — prior year flags don't carry
                acct.get("sort_order"),
                now, now,
            ),
        )

    # ── Prior-year balances ───────────────────────────────────────────────
    # Write PY final/ftax for every rolled account so Report tab can show
    # comparative columns immediately after rollforward.
    for acct in data.get("accounts", []):
        old_id = acct["account_id"]
        new_id = old_to_new.get(old_id)
        if not new_id:
            continue
        bal = balances.get(old_id, {"final": 0.0, "ftax": 0.0})
        conn.execute(
            """INSERT OR REPLACE INTO prior_year_balances
                 (py_balance_id, job_id, account_id, py_final_balance,
                  py_ftax_balance, source, entered_at)
               VALUES (?, ?, ?, ?, ?, 'rollforward', ?)""",
            (new_uuid(), new_job_id, new_id,
             bal["final"], bal["ftax"], now),
        )

    # ── Tax lines (seed if absent) ────────────────────────────────────────
    # Tax lines are global (no job_id). Copy any that don't already exist.
    for tl in data.get("tax_lines", []):
        conn.execute(
            """INSERT OR IGNORE INTO tax_lines
                 (tax_line_id, entity_type, financial_statement, section,
                  section_sort_order, line_code, line_name, sort_order,
                  is_active, tax_year)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (
                tl["tax_line_id"], tl["entity_type"],
                tl["financial_statement"],
                tl.get("section", ""),
                tl.get("section_sort_order", 0),
                tl["line_code"], tl["line_name"],
                tl.get("sort_order", 0),
                tl.get("is_active", 1),
                tl.get("tax_year"),
            ),
        )

    # ── Mappings ──────────────────────────────────────────────────────────
    for m in data.get("mappings", []):
        old_aid = m["account_id"]
        new_aid = old_to_new.get(old_aid)
        if not new_aid:
            continue
        conn.execute(
            """INSERT OR IGNORE INTO mappings
                 (mapping_id, account_id, job_id, tax_line_id, section_id,
                  mapped_by, mapped_at, notes)
               VALUES (?,?,?,?,?,?,?,?)""",
            (
                new_uuid(), new_aid, new_job_id,
                m.get("tax_line_id"),
                m.get("section_id"),
                performed_by,
                now,
                "Rolled forward from prior year",
            ),
        )

    # ── Account groups ────────────────────────────────────────────────────
    # old_group_id → new_group_id
    old_grp_to_new: dict[str, str] = {}
    groups = data.get("account_groups", [])

    # Insert in sort_order (parents before children — same guarantee as export)
    for g in sorted(groups, key=lambda x: x.get("sort_order", 0)):
        new_gid = new_uuid()
        old_grp_to_new[g["group_id"]] = new_gid
        new_parent = old_grp_to_new.get(g.get("parent_id") or "")
        conn.execute(
            """INSERT OR IGNORE INTO account_groups
                 (group_id, job_id, name, parent_id, sort_order, created_at)
               VALUES (?,?,?,?,?,?)""",
            (
                new_gid, new_job_id, g["name"],
                new_parent,
                g.get("sort_order", 0),
                now,
            ),
        )

    # ── Account group members ─────────────────────────────────────────────
    for m in data.get("account_group_members", []):
        new_gid = old_grp_to_new.get(m["group_id"])
        new_aid = old_to_new.get(m["account_id"])
        if new_gid and new_aid:
            conn.execute(
                "INSERT OR IGNORE INTO account_group_members "
                "(group_id, account_id) VALUES (?,?)",
                (new_gid, new_aid),
            )

    # ── Activity log ──────────────────────────────────────────────────────
    log_activity(
        conn,
        job_id      = new_job_id,
        event_type  = "rollforward_created",
        entity_type = "job",
        entity_id   = prior_job_id,
        description = f"Rolled forward from prior year job {prior_job_id[:8]}…",
        performed_by= performed_by,
    )
