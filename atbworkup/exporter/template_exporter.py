"""
Tax line template Excel exporter.

Exports existing templates from the settings DB to .xlsx files that can be
re-imported via template_importer.import_template_from_excel().

Also produces a blank template scaffold for creating new custom templates.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Brand constants — hex strings so this module has no Qt dependency
_NAVY_HEX  = "1A2B4C"
_WHITE_HEX = "FFFFFF"

# Column layout for content sheets
_HEADERS = ["section_sort_order", "section", "line_code", "line_name", "sort_order"]
_COL_WIDTHS = {
    1: 18,   # section_sort_order
    2: 28,   # section
    3: 16,   # line_code
    4: 40,   # line_name
    5: 18,   # sort_order
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _header_font() -> Font:
    return Font(name="Calibri", bold=True, color=_WHITE_HEX, size=11)


def _header_fill() -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=_NAVY_HEX)


def _header_alignment() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def _apply_header_row(ws, headers: list[str]) -> None:
    """Write a styled header row to row 1 of the worksheet."""
    fill  = _header_fill()
    font  = _header_font()
    align = _header_alignment()
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font  = font
        cell.fill  = fill
        cell.alignment = align


def _apply_column_widths(ws) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _freeze_header(ws) -> None:
    ws.freeze_panes = "A2"


def _write_info_sheet(wb, template_name: str, entity_type_code: str, description: str) -> None:
    """Write (or overwrite) the Info sheet in wb."""
    if "Info" in wb.sheetnames:
        del wb["Info"]
    ws = wb.create_sheet("Info", 0)  # insert at front

    bold_font  = Font(name="Calibri", bold=True, size=11)
    label_fill = PatternFill(fill_type="solid", fgColor="D0D8E8")

    meta = [
        ("Template Name",    template_name),
        ("Entity Type Code", entity_type_code),
        ("Description",      description),
    ]
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 44

    for row_idx, (label, value) in enumerate(meta, start=2):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font  = bold_font
        label_cell.fill  = label_fill
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.alignment = Alignment(horizontal="left")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_template_to_excel(entity_type_code: str, output_path: str | Path) -> None:
    """
    Export an existing template from the settings DB to an Excel file.

    Creates:
      - An "Info" sheet with template metadata
      - One sheet per unique financial_statement value, styled with navy headers,
        frozen first row, and calibrated column widths.

    Raises
    ------
    ValueError  if entity_type_code is not found in the settings DB.
    """
    from atbworkup.db.settings import settings_connection

    output_path = Path(output_path)

    with settings_connection() as conn:
        # Detect optional columns
        tlt_cols = {row[1] for row in conn.execute("PRAGMA table_info(tax_line_templates)")}
        has_template_name = "template_name" in tlt_cols
        has_is_builtin    = "is_builtin" in tlt_cols

        rows = conn.execute(
            "SELECT * FROM tax_line_templates "
            "WHERE entity_type = ? "
            "ORDER BY financial_statement, section_sort_order, sort_order",
            (entity_type_code,),
        ).fetchall()

        if not rows:
            raise ValueError(f"No template found for entity type code '{entity_type_code}'")

        # Pull display name from first row when available
        first = dict(rows[0])
        if has_template_name and first.get("template_name"):
            template_name = first["template_name"]
        else:
            template_name = entity_type_code

        description = ""  # not stored in the current schema

    # Group rows by financial_statement
    from collections import defaultdict
    by_fs: dict[str, list] = defaultdict(list)
    for row in rows:
        by_fs[dict(row)["financial_statement"]].append(dict(row))

    wb = openpyxl.Workbook()
    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Info sheet first
    _write_info_sheet(wb, template_name, entity_type_code, description)

    # One content sheet per financial_statement
    for fs_name, fs_rows in by_fs.items():
        ws = wb.create_sheet(title=fs_name)
        _apply_header_row(ws, _HEADERS)
        _apply_column_widths(ws)
        _freeze_header(ws)

        for row_idx, row_dict in enumerate(fs_rows, start=2):
            ws.cell(row=row_idx, column=1, value=row_dict.get("section_sort_order"))
            ws.cell(row=row_idx, column=2, value=row_dict.get("section", ""))
            ws.cell(row=row_idx, column=3, value=row_dict.get("line_code", ""))
            ws.cell(row=row_idx, column=4, value=row_dict.get("line_name", ""))
            ws.cell(row=row_idx, column=5, value=row_dict.get("sort_order"))

    wb.save(str(output_path))


def export_blank_template(output_path: str | Path) -> None:
    """
    Export a blank template scaffold with headers and clearly labelled example rows.

    Creates:
      - "Info" sheet with placeholder metadata
      - "Balance Sheet" sheet with two example rows
      - "Profit & Loss" sheet with two example rows

    The example rows use line_codes like "EXAMPLE-1" and should be deleted
    before re-importing.
    """
    output_path = Path(output_path)

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Info sheet
    _write_info_sheet(
        wb,
        template_name="My Custom Template",
        entity_type_code="MyCustomCode",
        description="Enter a description of this template",
    )

    # Example data per sheet
    example_sheets = {
        "Balance Sheet": [
            (10, "Assets",      "EXAMPLE-1", "DELETE THIS ROW — example only", 10),
            (10, "Assets",      "EXAMPLE-2", "DELETE THIS ROW — example only", 20),
            (20, "Liabilities", "EXAMPLE-3", "DELETE THIS ROW — example only", 10),
            (30, "Equity",      "EXAMPLE-4", "DELETE THIS ROW — example only", 10),
        ],
        "Profit & Loss": [
            (10, "Revenue",  "EXAMPLE-1", "DELETE THIS ROW — example only", 10),
            (10, "Revenue",  "EXAMPLE-2", "DELETE THIS ROW — example only", 20),
            (20, "Expenses", "EXAMPLE-3", "DELETE THIS ROW — example only", 10),
            (20, "Expenses", "EXAMPLE-4", "DELETE THIS ROW — example only", 20),
        ],
    }

    example_font = Font(name="Calibri", italic=True, color="888888", size=10)

    for sheet_name, example_rows in example_sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        _apply_header_row(ws, _HEADERS)
        _apply_column_widths(ws)
        _freeze_header(ws)

        for row_idx, (sec_sort, section, line_code, line_name, sort_ord) in enumerate(
            example_rows, start=2
        ):
            for col_idx, value in enumerate(
                [sec_sort, section, line_code, line_name, sort_ord], start=1
            ):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = example_font

    wb.save(str(output_path))
