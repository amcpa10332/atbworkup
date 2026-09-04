"""
Review package exporter — produces a `.atbr.xlsx` file.

Tab order (all visible unless noted):
  1. Cover            — client/job info, preparer/reviewer, what's in the package
  2. Balance Sheet    — SUMMARY: one row per tax line with section subtotals + current period NI
  3. Income Statement — SUMMARY: one row per tax line with section subtotals + net income
  4. Tax Grouping     — DETAIL: individual accounts nested under section → tax line (BS then IS)
  5. Trial Balance    — flat working detail with all 7 columns for all accounts
  6. Journal Entries  — one row per JE line, grouped by entry number
  7. Notes            — preparer notes

Hidden (xlSheetVeryHidden):
  __data             — machine-readable JSON snapshot for re-import / lineage
  __manifest         — job metadata + SHA-256 of __data
"""
from __future__ import annotations

import datetime
import hashlib
import json
import re
import uuid
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from atbworkup.db.connection import db_connection
from atbworkup.models.accounts import get_grouped_balances
from atbworkup.models.journal_entries import get_entries, get_lines
from atbworkup.models.notes import get_notes
from atbworkup.data.tax_line_categories import (
    CATEGORY_REVENUE, CATEGORY_COGS, CATEGORY_EQUITY, CATEGORY_CURRENT_ASSET,
    CATEGORY_OPEX, CATEGORY_DISTRIBUTION, CATEGORY_SCHEDULE_K,
    ASSET_CATEGORIES, LIABILITY_CATEGORIES,
    ASSET_BUCKET_ORDER, LIABILITY_BUCKET_ORDER,
)


# NOTE: Sheet/workbook password protection was removed — some Excel builds
# treated the protected/hidden combination as a corrupt file and refused to
# open it. The __manifest SHA-256 checksum is the tamper detector; the __data
# and __manifest tabs stay veryHidden (invisible in the UI) without any
# password, which is enough to keep normal users out of them.

# ── Brand colours ─────────────────────────────────────────────────────────────
_NAVY   = "1A2B4C"
_WHITE  = "FFFFFF"
_PLAT   = "E5E5E5"
_LIGHT  = "D0D8E8"
_BLACK  = "000000"

# ── Amount column keys and display headers ────────────────────────────────────
_AMT_KEYS = ["pbc_balance", "aje", "adj", "rje", "final", "ftje", "ftax"]
_AMT_HDRS = ["PBC",         "AJE", "ADJ", "RJE", "FINAL", "FTJE", "FTAX"]

_FS_LABELS = {"BalanceSheet": "Balance Sheet", "ProfitAndLoss": "Income Statement"}
_FS_ORDER  = ["BalanceSheet", "ProfitAndLoss", "Unmapped"]

_ENTITY_LABELS = {
    "1120S": "S-Corporation (1120S)",
    "1065":  "Partnership (1065)",
    "1120":  "C-Corporation (1120)",
    "ScheduleC": "Sole Proprietor (Sch C)",
    "990":   "Non-Profit (990)",
    "1041":  "Trust / Estate (1041)",
}


# ── Public API ────────────────────────────────────────────────────────────────

def suggested_filename(job: dict, version: int | None = None) -> str:
    from atbworkup.utils.naming import suggested_filename as _sfn
    v = version if version is not None else job.get("workflow_version", 1)
    return _sfn(
        tax_year=job["tax_year"],
        client_name=job["client_name"],
        status=job.get("status", "Preparation in Progress"),
        version=v,
    )


def next_version(conn, job_id: str) -> int:
    """
    Next package version for this job, based on the packages table itself —
    NOT job.workflow_version, which tracks workflow-status transitions
    (Preparation -> Review -> ...) and advances independently of how many
    times a package has actually been exported.
    """
    row = conn.execute(
        "SELECT COALESCE(MAX(version_number), 0) FROM packages WHERE job_id = ?",
        (job_id,),
    ).fetchone()
    return (row[0] or 0) + 1


def _atomic_save(wb, output_path: Path) -> None:
    """
    Write the workbook to a temp file in the same directory, then atomically
    replace the target. openpyxl's wb.save() writes directly to the target
    path — if that write is interrupted (a second save firing on top of it,
    a sync client or antivirus grabbing the file mid-write, a crash), the
    result is a zip with a corrupted central directory: readable-looking but
    unopenable in Excel or this app. Writing to a temp file first means a
    failed/interrupted write never touches the real file; os.replace() is
    atomic on the same volume, so a second concurrent save can't interleave
    with this one either.
    """
    import os
    import tempfile
    output_path = Path(output_path)
    fd, tmp_name = tempfile.mkstemp(
        prefix=f".{output_path.stem}.", suffix=".tmp", dir=str(output_path.parent)
    )
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        wb.save(str(tmp_path))
        os.replace(str(tmp_path), str(output_path))
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise


def save_workup(conn, *, job: dict, output_path: str | Path, performed_by: str) -> None:
    """
    Write current binder state to output_path without incrementing the version
    or updating job status.  Used for auto-save / manual Save operations.
    """
    output_path = Path(output_path)
    version = job.get("workflow_version", 1)
    wb, _   = _build_workbook(conn, job, version, performed_by, output_path.name, _now())
    _atomic_save(wb, output_path)


def transition_workup_status(
    conn,
    *,
    job: dict,
    new_status: str,
    output_dir: Path,
    performed_by: str,
) -> tuple[dict, Path]:
    """
    Transition the job to new_status, increment workflow_version, build a new
    .atbr.xlsx in output_dir with the new name, and return (updated_job, new_path).
    """
    from atbworkup.models.job import transition_status, get_job
    from atbworkup.utils.naming import suggested_filename as _sfn
    from atbworkup.db.connection import db_connection as _dbc

    new_version = transition_status(conn, job["job_id"], new_status, performed_by)
    # Re-read the updated job row
    updated_job = dict(conn.execute(
        "SELECT * FROM job WHERE job_id = ?", (job["job_id"],)
    ).fetchone())

    filename = _sfn(
        tax_year=updated_job["tax_year"],
        client_name=updated_job["client_name"],
        status=new_status,
        version=new_version,
    )
    new_path = output_dir / filename
    wb, _ = _build_workbook(
        conn, updated_job, new_version, performed_by, filename, _now()
    )
    _atomic_save(wb, new_path)
    return updated_job, new_path


def export_review_package(conn, *, job: dict, output_path: str | Path,
                          performed_by: str) -> dict:
    """Build and write the .atbr.xlsx file. Returns the inserted packages row."""
    output_path = Path(output_path)
    version     = next_version(conn, job["job_id"])
    now_str     = _now()

    wb, data_content = _build_workbook(
        conn, job, version, performed_by, output_path.name, now_str
    )
    checksum = _sha256(data_content)

    _atomic_save(wb, output_path)

    # ── Persist package record ────────────────────────────────────────────
    pkg_id = uuid.uuid4().hex
    prior  = conn.execute(
        "SELECT package_id FROM packages WHERE job_id = ? ORDER BY version_number DESC LIMIT 1",
        (job["job_id"],),
    ).fetchone()

    conn.execute(
        """INSERT INTO packages
               (package_id, job_id, version_number, package_type, status_label,
                file_name, file_path, exported_by, exported_at,
                prior_package_id, checksum)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
        (pkg_id, job["job_id"], version, "review", "Ready for Review",
         output_path.name, str(output_path), performed_by, now_str,
         prior["package_id"] if prior else None, checksum),
    )
    conn.execute(
        "UPDATE job SET status = 'Ready for Review', updated_at = ? WHERE job_id = ?",
        (now_str, job["job_id"]),
    )

    from atbworkup.models.activity import log_activity
    log_activity(
        conn,
        job_id=job["job_id"],
        event_type="exported_package",
        description=f"Exported review package V{version:02d}: {output_path.name}",
        performed_by=performed_by,
    )

    return conn.execute(
        "SELECT * FROM packages WHERE package_id = ?", (pkg_id,)
    ).fetchone()


# ── Shared workbook builder ───────────────────────────────────────────────────

def _build_workbook(conn, job: dict, version: int, performed_by: str,
                    filename: str, now_str: str):
    """
    Construct the full .atbr.xlsx openpyxl Workbook in memory.
    Returns (wb, data_content_str).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_cover_tab(wb, job, version, performed_by, now_str)
    if job.get("entity_type") == "Consolidated":
        # A consolidated job has no accounts of its own — every account
        # lives in a subsidiary's own file — so the regular tab writers
        # (which query THIS job's accounts table) would always render empty.
        # Build the report tabs from the same combined-subsidiary data the
        # interactive Consolidation window shows instead.
        _write_consolidated_report_tabs(wb, conn, job)
    else:
        _write_bs_summary_tab(wb, conn, job)
        _write_is_summary_tab(wb, conn, job)
        _write_tax_grouping_tab(wb, conn, job)
        _write_trial_balance_tab(wb, conn, job)
        _write_journal_entries_tab(wb, conn, job)
    _write_notes_tab(wb, conn, job)

    data_content = _write_data_tab(wb, conn, job)
    checksum     = _sha256(data_content)
    _write_manifest_tab(wb, conn, job, version, performed_by, checksum, filename, now_str)

    for ws in wb.worksheets:
        if ws.title.startswith("__"):
            ws.sheet_state = "veryHidden"
    return wb, data_content


# ── Consolidated report tabs ───────────────────────────────────────────────────
# A consolidated job holds no accounts itself; every real account lives in a
# subsidiary's own .atbr.xlsx. Read and combine them the same way the
# interactive Consolidation window does (models/consolidation_read.py), then
# write that combined data into real Excel tabs instead of the always-empty
# regular ones.

def _write_consolidated_report_tabs(wb, conn, job: dict) -> None:
    from atbworkup.models import consolidation_calc
    calc = consolidation_calc.compute_combined(conn, job)

    _write_consolidated_statement_tab(
        wb, "Combined Balance Sheet", calc["combined_bs"],
        calc["elim_by_section"], calc["cte_by_section"], is_bs=True,
        net_income=calc["combined_net_income"],
    )
    _write_consolidated_statement_tab(
        wb, "Combined Income Statement", calc["combined_pl"],
        calc["elim_by_section"], calc["cte_by_section"], is_bs=False,
        pl_grand_override=(calc["pl_base_ni"], calc["pl_elim_ni"], calc["pl_cte_ni"]),
        sch_k_override=calc["sch_k_total_ni"] if calc["has_sch_k"] else None,
    )
    # detail_bs/detail_pl are built in lockstep with member_labels in the
    # same subsidiary loop, so index-aligning them here is safe.
    member_detail = list(zip(calc["member_labels"], calc["detail_bs"], calc["detail_pl"]))
    _write_consolidated_tax_grouping_tab(wb, [
        (label, bs_lbl_detail[1], pl_lbl_detail[1])
        for label, bs_lbl_detail, pl_lbl_detail in member_detail
    ])
    if calc["errors"]:
        _write_consolidated_errors_tab(wb, calc["errors"])


def _sec_sort(lines: dict) -> int:
    return min(v["sort"] for v in lines.values()) // 1000 if lines else 999


def _sec_match_export(section_name: str, by_section_map: dict) -> float:
    """Exact (trimmed, case-insensitive) match — see the same helper in
    consolidation_window.py for why prefix matching is unsafe here."""
    sl = section_name.strip().lower()
    return sum(amt for key, amt in by_section_map.items()
              if key.strip() and key.strip().lower() == sl)


def _write_consolidated_section_group(ws, sections_subset: dict, elim_by_section: dict,
                                      cte_by_section: dict, r: int) -> tuple[int, dict]:
    """Write section headers/lines/subtotals for a subset of sections (one
    GAAP bucket's worth). Returns (next_row, {combined, elim, cte} totals)."""
    totals = {"combined": 0.0, "elim": 0.0, "cte": 0.0}
    for section, lines in sorted(sections_subset.items(), key=lambda kv: _sec_sort(kv[1])):
        sec_final = sum(v["final"] for v in lines.values())
        sec_elim  = _sec_match_export(section, elim_by_section)
        sec_cte   = _sec_match_export(section, cte_by_section)
        sec_net   = sec_final + sec_elim
        sec_tax   = sec_net + sec_cte

        _write_summary_section_header(ws, section, r, ncols=6)
        r += 1
        for lname, vals in sorted(lines.items(), key=lambda x: x[1]["sort"]):
            _write_summary_line_row(
                ws, [f"  {lname}", vals["final"] or None, None, vals["final"] or None, None, vals["final"] or None],
                r,
            )
            r += 1
        _write_total_row(
            ws, [f"Total {section}", sec_final or None,
                 sec_elim or None, sec_net or None, sec_cte or None, sec_tax or None],
            r, amount_start_col=2,
        )
        r += 2

        totals["combined"] += sec_final
        totals["elim"]     += sec_elim
        totals["cte"]      += sec_cte
    return r, totals


def _write_consolidated_statement_tab(wb, title: str, data: dict,
                                      elim_by_section: dict, cte_by_section: dict,
                                      is_bs: bool, net_income: float = 0.0,
                                      pl_grand_override: tuple[float, float, float] | None = None,
                                      sch_k_override: float | None = None):
    ws = wb.create_sheet(title)
    ws.column_dimensions["A"].width = 42
    for col in "BCDEF":
        ws.column_dimensions[col].width = 16
    headers = ["Section / Tax Line", "Combined FINAL", "EJEs", "Net Consolidated",
               "Tax Entries (CTE)", "Net Tax"]
    _write_header_row(ws, headers, 1)

    r = 2

    if is_bs:
        # Presentation-ready Balance Sheet: sections bucketed into GAAP
        # subtotals by category, same as the single-entity tab. No balance-
        # check row here by design — that lives in the TB view, not a report
        # handed to a bank or other outside reader.
        known_categories = ASSET_CATEGORIES | LIABILITY_CATEGORIES | {CATEGORY_EQUITY}

        def _sections_with(categories: set, catch_leftovers: bool = False) -> dict:
            out = {}
            for section, lines in data.items():
                cat = next(iter(lines.values()), {}).get("category", "")
                if cat in categories or (catch_leftovers and cat not in known_categories):
                    out[section] = lines
            return out

        # Current Assets stands alone; Fixed + Other Long-Term Assets roll up
        # into a "Total Noncurrent Assets" subtotal before Total Assets.
        asset_totals = {"combined": 0.0, "elim": 0.0, "cte": 0.0}
        noncurrent_totals = {"combined": 0.0, "elim": 0.0, "cte": 0.0}
        for i, (category, bucket_label) in enumerate(ASSET_BUCKET_ORDER):
            is_last = i == len(ASSET_BUCKET_ORDER) - 1
            subset = _sections_with({category}, catch_leftovers=is_last)
            if not subset:
                continue
            r, bucket_totals = _write_consolidated_section_group(ws, subset, elim_by_section, cte_by_section, r)
            bt_net = bucket_totals["combined"] + bucket_totals["elim"]
            bt_tax = bt_net + bucket_totals["cte"]
            _write_total_row(
                ws, [bucket_label, bucket_totals["combined"] or None, bucket_totals["elim"] or None,
                     bt_net or None, bucket_totals["cte"] or None, bt_tax or None],
                r, amount_start_col=2,
            )
            r += 2
            for k in asset_totals:
                asset_totals[k] += bucket_totals[k]
                if category != CATEGORY_CURRENT_ASSET:
                    noncurrent_totals[k] += bucket_totals[k]
        if any(noncurrent_totals.values()):
            ncn = noncurrent_totals["combined"] + noncurrent_totals["elim"]
            nct = ncn + noncurrent_totals["cte"]
            _write_total_row(
                ws, ["Total Noncurrent Assets", noncurrent_totals["combined"] or None, noncurrent_totals["elim"] or None,
                     ncn or None, noncurrent_totals["cte"] or None, nct or None],
                r, amount_start_col=2,
            )
            r += 2
        an = asset_totals["combined"] + asset_totals["elim"]
        at = an + asset_totals["cte"]
        _write_total_row(
            ws, ["TOTAL ASSETS", asset_totals["combined"] or None, asset_totals["elim"] or None,
                 an or None, asset_totals["cte"] or None, at or None],
            r, amount_start_col=2,
        )
        r += 2

        liab_totals = {"combined": 0.0, "elim": 0.0, "cte": 0.0}
        for category, bucket_label in LIABILITY_BUCKET_ORDER:
            subset = _sections_with({category})
            if not subset:
                continue
            r, bucket_totals = _write_consolidated_section_group(ws, subset, elim_by_section, cte_by_section, r)
            bt_net = bucket_totals["combined"] + bucket_totals["elim"]
            bt_tax = bt_net + bucket_totals["cte"]
            _write_total_row(
                ws, [bucket_label, bucket_totals["combined"] or None, bucket_totals["elim"] or None,
                     bt_net or None, bucket_totals["cte"] or None, bt_tax or None],
                r, amount_start_col=2,
            )
            r += 2
            for k in liab_totals:
                liab_totals[k] += bucket_totals[k]
        ln_ = liab_totals["combined"] + liab_totals["elim"]
        lt = ln_ + liab_totals["cte"]
        _write_total_row(
            ws, ["Total Liabilities", liab_totals["combined"] or None, liab_totals["elim"] or None,
                 ln_ or None, liab_totals["cte"] or None, lt or None],
            r, amount_start_col=2,
        )
        r += 2

        equity_subset = _sections_with({CATEGORY_EQUITY})
        equity_totals = {"combined": 0.0, "elim": 0.0, "cte": 0.0}
        if equity_subset:
            r, equity_totals = _write_consolidated_section_group(ws, equity_subset, elim_by_section, cte_by_section, r)
        if abs(net_income) >= 0.005:
            _write_summary_line_row(ws, ["Current Period Net Income (from P&L)", None, None, None, None, net_income], r)
            r += 1
        equity_grand = equity_totals["combined"] + equity_totals["elim"] + net_income
        equity_tax   = equity_grand + equity_totals["cte"]
        _write_total_row(
            ws, ["Total Equity", equity_totals["combined"] or None, equity_totals["elim"] or None,
                 equity_grand or None, equity_totals["cte"] or None, equity_tax or None],
            r, amount_start_col=2,
        )
        r += 2

        liab_eq_net = ln_ + equity_grand
        liab_eq_tax = lt + equity_tax
        _write_total_row(
            ws, ["TOTAL LIABILITIES & EQUITY", (liab_totals["combined"] + equity_totals["combined"]) or None,
                 (liab_totals["elim"] + equity_totals["elim"]) or None, liab_eq_net or None,
                 (liab_totals["cte"] + equity_totals["cte"]) or None, liab_eq_tax or None],
            r, amount_start_col=2,
        )
    else:
        grand_combined = 0.0
        grand_elim = grand_cte = 0.0
        rev_total = cogs_total = 0.0
        gross_profit_written = False
        seen_revenue = False

        for section, lines in sorted(data.items(), key=lambda kv: _sec_sort(kv[1])):
            sec_final = sum(v["final"] for v in lines.values())
            sec_elim  = _sec_match_export(section, elim_by_section)
            sec_cte   = _sec_match_export(section, cte_by_section)
            sec_net   = sec_final + sec_elim
            sec_tax   = sec_net + sec_cte
            category  = next(iter(lines.values()), {}).get("category", "")

            _write_summary_section_header(ws, section, r, ncols=6)
            r += 1
            for lname, vals in sorted(lines.items(), key=lambda x: x[1]["sort"]):
                _write_summary_line_row(
                    ws, [f"  {lname}", vals["final"] or None, None, vals["final"] or None, None, vals["final"] or None],
                    r,
                )
                r += 1
            _write_total_row(
                ws, [f"Total {section}", sec_final or None,
                     sec_elim or None, sec_net or None, sec_cte or None, sec_tax or None],
                r, amount_start_col=2,
            )
            r += 2

            grand_combined += sec_final
            grand_elim     += sec_elim
            grand_cte      += sec_cte

            if category == CATEGORY_REVENUE:
                rev_total += sec_final
                seen_revenue = True
            elif category == CATEGORY_COGS:
                cogs_total += sec_final
                if not gross_profit_written and seen_revenue:
                    _write_total_row(ws, ["Gross Profit", rev_total - cogs_total],
                                     r, amount_start_col=2)
                    r += 2
                    gross_profit_written = True

        if pl_grand_override is not None:
            # Revenue/expense-type sections are both displayed positive, so
            # summing them (grand_combined above) overstates net income. Use
            # the correctly-signed figure from consolidation_calc instead.
            grand_combined, grand_elim, grand_cte = pl_grand_override

        if sch_k_override is not None and abs(sch_k_override) >= 0.005:
            obi = grand_combined + grand_elim + grand_cte - sch_k_override
            _write_total_row(ws, ["Ordinary Business Income / (Loss)", obi], r, amount_start_col=2)
            r += 1
            _write_total_row(ws, ["+ Schedule K Items (Net)", sch_k_override], r, amount_start_col=2)
            r += 1
        grand_net = grand_combined + grand_elim
        grand_tax = grand_net + grand_cte
        _write_total_row(
            ws, ["Net Income / (Loss)", grand_combined or None,
                 grand_elim or None, grand_net or None, grand_cte or None, grand_tax or None],
            r, amount_start_col=2,
        )
    ws.freeze_panes = "A2"


def _write_consolidated_tax_grouping_tab(wb, member_rows: list[tuple[str, dict, dict]]):
    """
    Account-level detail across every subsidiary, in one tab — lets a
    preparer filter/pivot on "Tax Line" (e.g. "Other Deductions") to see
    every contributing account across all subs at once.
    """
    ws = wb.create_sheet("Combined Tax Grouping")
    headers = ["Subsidiary", "Statement", "Section", "Tax Line", "Acct #", "Account Name", "FINAL"]
    _write_header_row(ws, headers, 1)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 32
    ws.column_dimensions["G"].width = 14

    r = 2
    for label, bs_detail, pl_detail in member_rows:
        for stmt_label, detail in (("Balance Sheet", bs_detail), ("Income Statement", pl_detail)):
            for section, lines in detail.items():
                for lname, accts in lines.items():
                    for acct in accts:
                        row_data = [
                            label, stmt_label, section, lname,
                            acct.get("number") or "", acct.get("name") or "",
                            acct.get("final") or None,
                        ]
                        _write_data_row(ws, row_data, r, amount_start_col=7)
                        r += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(r - 1, 1)}"


def _write_consolidated_errors_tab(wb, errors: list[str]):
    ws = wb.create_sheet("Combined — Warnings")
    ws.column_dimensions["A"].width = 90
    ws.cell(row=1, column=1, value="Some subsidiary binders could not be read for this export:").font = Font(bold=True, color="CC0000")
    for i, err in enumerate(errors, start=2):
        ws.cell(row=i, column=1, value=f"• {err}")


# ── Cover tab ─────────────────────────────────────────────────────────────────

def _write_cover_tab(wb, job: dict, version: int, performed_by: str, now_str: str):
    ws = wb.create_sheet("Cover")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 48

    navy_fill = _fill(_NAVY)
    plat_fill = _fill(_PLAT)
    light_fill = _fill(_LIGHT)

    def _hdr(row, text):
        c = ws.cell(row=row, column=1, value=text)
        c.font      = Font(bold=True, color=_WHITE, size=12)
        c.fill      = navy_fill
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 20

    def _row(row, label, value, shade=False):
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        f  = plat_fill if shade else None
        lc.font = Font(bold=True, size=10, color=_NAVY)
        vc.font = Font(size=10)
        if shade:
            lc.fill = f
            vc.fill = f
        ws.row_dimensions[row].height = 16

    r = 1
    _hdr(r, "Trial Balance Review Package")
    r += 1
    ws.row_dimensions[r].height = 8
    r += 1

    _hdr(r, "Client & Job Information")
    r += 1
    fields = [
        ("Client Name",       job.get("client_name")),
        ("Entity Name",       job.get("entity_name")),
        ("Entity Type",       _ENTITY_LABELS.get(job.get("entity_type",""), job.get("entity_type",""))),
        ("Tax Year",          job.get("tax_year")),
        ("Accounting System", job.get("accounting_system")),
    ]
    for i, (lbl, val) in enumerate(fields):
        _row(r, lbl, val or "—", shade=(i % 2 == 1))
        r += 1

    r += 1
    _hdr(r, "Preparer & Reviewer")
    r += 1
    prep_fields = [
        ("Prepared By", job.get("prepared_by")),
        ("Reviewer",    job.get("reviewer")),
        ("Status",      job.get("status")),
    ]
    for i, (lbl, val) in enumerate(prep_fields):
        _row(r, lbl, val or "—", shade=(i % 2 == 1))
        r += 1

    r += 1
    _hdr(r, "Package Information")
    r += 1
    pkg_fields = [
        ("Package Version",  f"V{version:02d}"),
        ("Exported By",      performed_by),
        ("Exported At",      now_str[:10]),
        ("File Format",      ".atbr.xlsx (ATBWorkup Review Package)"),
    ]
    for i, (lbl, val) in enumerate(pkg_fields):
        _row(r, lbl, val, shade=(i % 2 == 1))
        r += 1

    r += 1
    _hdr(r, "Contents of This Package")
    r += 1
    toc = [
        ("Balance Sheet",    "Summary balance sheet — one row per tax line with section subtotals. Equity section includes current period net income."),
        ("Income Statement", "Summary income statement — one row per tax line with section subtotals and net income total."),
        ("Tax Grouping",     "Detail report — individual accounts nested under section and tax line for both the balance sheet and income statement."),
        ("Trial Balance",    "Full working trial balance — every account with all seven adjustment columns: PBC, AJE, ADJ, RJE, FINAL, FTJE, FTAX."),
        ("Journal Entries",  "All adjusting, reclassifying, and federal tax journal entries with line-level detail."),
        ("Notes",            "Preparer notes linked to specific accounts or journal entries."),
    ]
    for i, (tab, desc) in enumerate(toc):
        lc = ws.cell(row=r, column=1, value=tab)
        vc = ws.cell(row=r, column=2, value=desc)
        f  = plat_fill if i % 2 == 1 else None
        lc.font = Font(bold=True, size=10, color=_NAVY)
        vc.font = Font(size=10)
        vc.alignment = Alignment(wrap_text=True)
        if f:
            lc.fill = f
            vc.fill = f
        ws.row_dimensions[r].height = 28
        r += 1


# ── Balance Sheet — summary tab ───────────────────────────────────────────────

def _write_bs_summary_tab(wb, conn, job: dict):
    """
    Presentation-ready, account-level Balance Sheet: no tax-line rollups —
    individual accounts (col B) under each section, section/bucket subtotals
    (col C), mid-level rollups (col D: Total Noncurrent Assets, Total
    Liabilities, Total Equity), and grand totals (col E: TOTAL ASSETS, TOTAL
    LIABILITIES & EQUITY) in the classic classified-statement staggered-
    column style. No balance-check row by design — that belongs on the
    Trial Balance, not a report handed to a bank or other outside reader.

    Liabilities and equity are stored DR=+/CR=- like everything else in this
    app, but a bank statement shows them as positive figures — so those two
    categories get sign-flipped for display here (assets, already debit-
    normal, are left alone so contra-asset lines like Accumulated
    Depreciation still net correctly against their parent asset).
    """
    ws = wb.create_sheet("Balance Sheet")
    groups = get_grouped_balances(conn, job["job_id"])
    bs_accounts = groups.get("BalanceSheet", [])
    pl_accounts = groups.get("ProfitAndLoss", [])
    # Net income = -sum(raw): revenue is credit-normal (raw negative),
    # expenses are debit-normal (raw positive); blindly summing raw adds
    # expenses instead of subtracting them from revenue.
    net_income = -round(sum(a.get("final", 0.0) for a in pl_accounts), 2)

    ws.column_dimensions["A"].width = 40
    for col in "BCDE":
        ws.column_dimensions[col].width = 16
    r = _write_report_title(ws, job, "Balance Sheet", ncols=5, as_of=True)

    known_categories = ASSET_CATEGORIES | LIABILITY_CATEGORIES | {CATEGORY_EQUITY}
    uncategorized = [a for a in bs_accounts if a.get("category") not in known_categories]

    # ── Assets ──
    # Current Assets stands alone; Fixed + Other Long-Term Assets roll up
    # into a "Total Noncurrent Assets" mid-level subtotal (col D) before the
    # Total Assets grand total (col E).
    asset_total = 0.0
    noncurrent_total = 0.0
    for category, _label in ASSET_BUCKET_ORDER:
        bucket_accounts = [a for a in bs_accounts if a.get("category") == category]
        if category == ASSET_BUCKET_ORDER[-1][0]:
            bucket_accounts += uncategorized
        if not bucket_accounts:
            continue
        r, bucket_total, _secs = _write_group_accounts(ws, bucket_accounts, r, sign=1.0)
        asset_total += bucket_total
        if category != CATEGORY_CURRENT_ASSET:
            noncurrent_total += bucket_total
    if noncurrent_total:
        _label_cell(ws, r, "Total Noncurrent Assets", bold=True)
        _amount_cell(ws, r, _COL_MID, noncurrent_total, bold=True, border=_SUBTOTAL_BORDER)
        r += 2
    _label_cell(ws, r, "TOTAL ASSETS", bold=True)
    _amount_cell(ws, r, _COL_GRAND, asset_total, bold=True, border=_GRAND_BORDER, always_show=True)
    r += 2

    # ── Liabilities ──
    liab_total = 0.0
    liab_sections: set = set()
    for category, _label in LIABILITY_BUCKET_ORDER:
        bucket_accounts = [a for a in bs_accounts if a.get("category") == category]
        if not bucket_accounts:
            continue
        r, bucket_total, secs = _write_group_accounts(ws, bucket_accounts, r, sign=-1.0)
        liab_total += bucket_total
        liab_sections |= secs
    if len(liab_sections) > 1:
        _label_cell(ws, r, "Total Liabilities", bold=True)
        _amount_cell(ws, r, _COL_MID, liab_total, bold=True, border=_SUBTOTAL_BORDER)
        r += 2

    # ── Equity (current period net income sits as a normal line item here,
    #    not a separate highlighted row — same accounts list, same styling). ──
    equity_accounts = [a for a in bs_accounts if a.get("category") == CATEGORY_EQUITY]
    equity_accounts = equity_accounts + [{
        "section": (equity_accounts[0]["section"] if equity_accounts else "Equity"),
        "account_name": "Current Period Net Income",
        "final": -net_income,  # stored sign convention: negate to match the -1 flip below
    }]
    r, equity_total, equity_sections = _write_group_accounts(ws, equity_accounts, r, sign=-1.0)
    if len(equity_sections) > 1:
        _label_cell(ws, r, "Total Equity", bold=True)
        _amount_cell(ws, r, _COL_MID, equity_total, bold=True, border=_SUBTOTAL_BORDER)
        r += 2

    liab_eq_total = round(liab_total + equity_total, 2)
    _label_cell(ws, r, "TOTAL LIABILITIES & EQUITY", bold=True)
    _amount_cell(ws, r, _COL_GRAND, liab_eq_total, bold=True, border=_GRAND_BORDER, always_show=True)
    ws.freeze_panes = "B4"


# ── Income Statement — summary tab ────────────────────────────────────────────

def _write_is_summary_tab(wb, conn, job: dict):
    """
    Presentation-ready, account-level Income Statement: individual accounts
    (col B) under each section, section subtotals (col C: Total Income,
    Total Cost of Goods Sold, Total Deductions, ...), Gross Profit as a
    mid-level rollup (col D), and Net Income / (Loss) as the grand total
    (col E) — same staggered-column style as the Balance Sheet.

    Revenue is stored DR=+/CR=- like everything else in this app (so it's
    raw negative); flipped here so it displays positive the way a bank
    expects, while COGS/expenses (already debit-normal, already positive)
    are left alone. Net income is -sum(raw) on the *unflipped* accounts —
    flipping revenue for display and then summing would double it against
    expenses instead of subtracting them.
    """
    ws = wb.create_sheet("Income Statement")
    groups = get_grouped_balances(conn, job["job_id"])
    pl_accounts = groups.get("ProfitAndLoss", [])
    net_income = -round(sum(a.get("final", 0.0) for a in pl_accounts), 2)

    ws.column_dimensions["A"].width = 40
    for col in "BCDE":
        ws.column_dimensions[col].width = 16
    r = _write_report_title(ws, job, "Income Statement", ncols=5, as_of=False)

    revenue_accounts = [a for a in pl_accounts if a.get("category") == CATEGORY_REVENUE]
    cogs_accounts    = [a for a in pl_accounts if a.get("category") == CATEGORY_COGS]
    other_accounts   = [a for a in pl_accounts
                        if a.get("category") not in (CATEGORY_REVENUE, CATEGORY_COGS)]

    revenue_total = cogs_total = 0.0
    if revenue_accounts:
        r, revenue_total, _secs = _write_group_accounts(ws, revenue_accounts, r, sign=-1.0)
    if cogs_accounts:
        r, cogs_total, _secs = _write_group_accounts(ws, cogs_accounts, r, sign=1.0)
        _label_cell(ws, r, "Gross Profit", bold=True)
        _amount_cell(ws, r, _COL_MID, revenue_total - cogs_total, bold=True, border=_SUBTOTAL_BORDER)
        r += 2
    if other_accounts:
        r, _total, _secs = _write_group_accounts(ws, other_accounts, r, sign=1.0)

    _label_cell(ws, r, "Net Income / (Loss)", bold=True)
    _amount_cell(ws, r, _COL_GRAND, net_income, bold=True, border=_GRAND_BORDER, always_show=True)
    ws.freeze_panes = "B4"


# ── Tax Grouping tab (account-level detail) ───────────────────────────────────

def _write_tax_grouping_tab(wb, conn, job: dict):
    """
    Both FS in one tab: Balance Sheet then Income Statement.
    Structure: FS banner → section header → tax line subheader → account rows
               → tax line subtotal → section total.

    Ties the return back to entries (objective #2): every account keeps its
    full PBC→FTAX column trail here, unlike the FINAL-only Balance Sheet /
    Income Statement presentation tabs. The Balance Sheet's Equity section
    also carries a synthetic "Current Period Net Income" line, same as the
    actual return, so this tab's Balance Sheet total ties to zero.
    """
    ws = wb.create_sheet("Tax Grouping")
    groups = get_grouped_balances(conn, job["job_id"])

    headers = ["Acct #", "Account Name", "Tax Line"] + _AMT_HDRS
    r = _write_report_title(ws, job, "Tax Grouping", ncols=len(headers), as_of=False)
    _write_header_row(ws, headers, row=r)
    r += 1
    _set_tb_col_widths(ws)

    for fs_key in ("BalanceSheet", "ProfitAndLoss"):
        accounts = groups.get(fs_key)
        if not accounts:
            continue
        fs_label = _FS_LABELS[fs_key]

        if fs_key == "BalanceSheet":
            # This tab stays in raw DR/CR convention throughout (unlike the
            # flipped-for-display BS/IS tabs), so the injected closing entry
            # must be the raw P&L sum itself, NOT its display-flipped
            # negative — that's the actual raw amount a closing entry would
            # credit to equity, and it's what makes this section (and the
            # "Total — Balance Sheet" grand total below it) tie to zero.
            pl_accounts = groups.get("ProfitAndLoss", [])
            net_income = {k: round(sum(a[k] for a in pl_accounts), 2) for k in _AMT_KEYS}
            equity_section = next(
                (a.get("section") for a in accounts if a.get("category") == CATEGORY_EQUITY),
                "Equity",
            )
            accounts = accounts + [{
                "account_number": "", "account_name": "Current Period Net Income",
                "section": equity_section, "section_sort_order": 999999,
                "line_name": "Current Period Net Income", "line_sort_order": 999999,
                "category": CATEGORY_EQUITY, **net_income,
            }]

        # FS banner (black background)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c, value=fs_label if c == 2 else "")
            cell.font = Font(bold=True, color=_WHITE, size=11)
            cell.fill = _fill(_BLACK)
        ws.row_dimensions[r].height = 20
        r += 1

        r, _ = _write_sectioned_accounts(ws, accounts, headers, r)

        # FS total
        totals = {k: round(sum(a[k] for a in accounts), 2) for k in _AMT_KEYS}
        total_row = ["", f"Total — {fs_label}", ""] + [
            (totals[k] if totals[k] != 0 else None) for k in _AMT_KEYS
        ]
        _write_total_row(ws, total_row, r, amount_start_col=4)
        r += 3

    ws.freeze_panes = "A5"


# ── Trial Balance tab ─────────────────────────────────────────────────────────

def _write_trial_balance_tab(wb, conn, job: dict):
    """
    Flat working detail, raw DR=+/CR=- convention throughout (this is the
    internal working tool, not a presentation copy). FS banners still group
    accounts visually, but there's exactly one grand total at the bottom —
    across every account in the binder — which is the actual balance check:
    it should equal 0.00 when the binder is in balance.
    """
    ws = wb.create_sheet("Trial Balance")
    groups = get_grouped_balances(conn, job["job_id"])

    headers = ["Acct #", "Account Name", "Tax Line"] + _AMT_HDRS
    r = _write_report_title(ws, job, "Trial Balance", ncols=len(headers), as_of=False)
    _write_header_row(ws, headers, row=r)
    r += 1
    _set_tb_col_widths(ws)

    grand_totals = {k: 0.0 for k in _AMT_KEYS}
    for fs_key in _FS_ORDER:
        accounts = groups.get(fs_key)
        if not accounts:
            continue
        fs_label = _FS_LABELS.get(fs_key, fs_key)

        # FS banner (visual grouping only — no subtotal row per statement)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c, value=fs_label if c == 2 else "")
            cell.font = Font(bold=True, color=_WHITE, size=11)
            cell.fill = _fill(_BLACK)
        ws.row_dimensions[r].height = 18
        r += 1

        for a in accounts:
            row_data = [a.get("account_number") or "", a["account_name"],
                        a.get("line_name") or ""]
            for key in _AMT_KEYS:
                v = a[key]
                row_data.append(v if v != 0 else None)
                grand_totals[key] += v
            _write_data_row(ws, row_data, r, amount_start_col=4)
            r += 1
        r += 1

    # Unlike every other total row, a zero here is the point (it's the
    # balance check) — show 0.00 explicitly instead of leaving it blank.
    total_row = ["", "Grand Total (should equal 0.00)", ""] + [
        round(grand_totals[k], 2) for k in _AMT_KEYS
    ]
    _write_total_row(ws, total_row, r, amount_start_col=4)
    for c in range(4, len(headers) + 1):
        ws.cell(row=r, column=c).border = _GRAND_BORDER

    ws.freeze_panes = "A5"


# ── Journal Entries tab ───────────────────────────────────────────────────────

_JE_TYPE_LABELS = {"AJE": "Adjusting Journal Entries", "RJE": "Reclassifying Journal Entries",
                   "FTJE": "Final Tax Journal Entries"}
_JE_TYPE_ORDER  = ["AJE", "RJE", "FTJE"]


def _write_journal_entries_tab(wb, conn, job: dict):
    """
    Entries grouped by type (AJE/RJE/FTJE), each in its own banded block with
    a subtotal; each entry within a block is visually separated (top border)
    with its own DR/CR subtotal, and the memo rendered in italic so it reads
    as a note rather than data. A grand total across all types closes the tab.
    """
    ws = wb.create_sheet("Journal Entries")
    entries = get_entries(conn, job["job_id"])
    by_type: dict[str, list] = {}
    for e in entries:
        by_type.setdefault(e["entry_type"], []).append(e)

    headers = ["Entry #", "Description", "Account #", "Account Name", "DR", "CR", "Memo"]
    r = _write_report_title(ws, job, "Journal Entries", ncols=len(headers), as_of=False)
    _write_header_row(ws, headers, row=r)
    r += 1
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 28

    grand_dr = grand_cr = 0.0
    types_present = [t for t in _JE_TYPE_ORDER if by_type.get(t)] + \
                    [t for t in by_type if t not in _JE_TYPE_ORDER]
    for etype in types_present:
        type_entries = by_type[etype]
        type_dr = type_cr = 0.0

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c, value=_JE_TYPE_LABELS.get(etype, etype) if c == 1 else "")
            cell.font = Font(bold=True, color=_WHITE, size=11)
            cell.fill = _fill(_BLACK)
        ws.row_dimensions[r].height = 20
        r += 1

        for entry in type_entries:
            lines = get_lines(conn, entry["aje_id"])
            entry_dr = entry_cr = 0.0
            entry_start = r
            for i, ln in enumerate(lines):
                dr = ln["amount"] if ln["amount"] > 0 else None
                cr = abs(ln["amount"]) if ln["amount"] < 0 else None
                entry_dr += dr or 0.0
                entry_cr += cr or 0.0
                row_data = [
                    entry["entry_number"] if i == 0 else "",
                    entry["description"]  if i == 0 else "",
                    ln.get("account_number") or "",
                    ln.get("account_name")  or "",
                    dr, cr,
                    ln.get("memo") or "",
                ]
                _write_data_row(ws, row_data, r, amount_start_col=5, amount_cols=2)
                if ln.get("memo"):
                    memo_cell = ws.cell(row=r, column=7)
                    f = memo_cell.font
                    memo_cell.font = Font(italic=True, size=f.sz, color="666666")
                # Divide each entry from its neighbor with a top border on
                # its first line, so a multi-line entry doesn't blur into
                # the next one.
                if i == 0:
                    for c in range(1, len(headers) + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.border = Border(top=Side(style="medium", color="999999"),
                                             left=cell.border.left, right=cell.border.right,
                                             bottom=cell.border.bottom)
                r += 1

            entry_total_row = ["", "", "", f"Entry {entry['entry_number']} Total",
                               round(entry_dr, 2) or None, round(entry_cr, 2) or None, ""]
            _write_total_row(ws, entry_total_row, r, amount_start_col=5)
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).border = _GRAND_BORDER
            r += 2
            type_dr += entry_dr
            type_cr += entry_cr

        type_total_row = ["", "", "", f"Total — {_JE_TYPE_LABELS.get(etype, etype)}",
                          round(type_dr, 2) or None, round(type_cr, 2) or None, ""]
        _write_total_row(ws, type_total_row, r, amount_start_col=5)
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = _GRAND_BORDER
        r += 2
        grand_dr += type_dr
        grand_cr += type_cr

    grand_row = ["", "", "", "GRAND TOTAL", round(grand_dr, 2) or None, round(grand_cr, 2) or None, ""]
    _write_total_row(ws, grand_row, r, amount_start_col=5)
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = _GRAND_BORDER

    ws.freeze_panes = "A5"


# ── Notes tab ─────────────────────────────────────────────────────────────────

def _write_notes_tab(wb, conn, job: dict):
    ws = wb.create_sheet("Notes")
    notes = get_notes(conn, job["job_id"], status_filter="All")

    headers = ["Type", "Linked To", "Note", "Created By", "Date", "Status"]
    _write_header_row(ws, headers, row=1)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12

    _GREY = "999999"
    for r, n in enumerate(notes, start=2):
        cleared = n["status"] == "Cleared"
        row_data = [
            n.get("note_type") or "preparer",
            n.get("linked_display") or "—",
            n["body"],
            n["created_by"],
            n["created_at"][:10],
            n["status"] if not cleared else f"Cleared by {n.get('cleared_by') or '—'}",
        ]
        for c, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(wrap_text=(c == 3), vertical="top")
            if cleared:
                cell.font = Font(color=_GREY, strike=True, size=10)
        ws.row_dimensions[r].height = 30

    ws.freeze_panes = "A2"


# ── Hidden tab: __data ────────────────────────────────────────────────────────

def _write_data_tab(wb, conn, job: dict) -> str:
    """
    Write full machine-readable snapshot (schema_version 2.0).
    Stores raw table rows so the binder can be fully reconstructed on re-open.
    """
    ws = wb.create_sheet("__data")

    raw_accounts = [
        dict(r) for r in conn.execute(
            "SELECT * FROM accounts WHERE job_id = ? ORDER BY sort_order",
            (job["job_id"],),
        )
    ]
    tax_lines = [
        dict(r) for r in conn.execute(
            "SELECT * FROM tax_lines WHERE entity_type = ? ORDER BY section_sort_order, sort_order",
            (job["entity_type"],),
        )
    ]
    mappings = [
        dict(r) for r in conn.execute(
            "SELECT * FROM mappings WHERE job_id = ?",
            (job["job_id"],),
        )
    ]
    entries = [
        dict(r) for r in conn.execute(
            "SELECT * FROM journal_entries WHERE job_id = ? ORDER BY entry_number",
            (job["job_id"],),
        )
    ]
    entry_lines = [
        dict(r) for r in conn.execute(
            """SELECT jel.*
               FROM journal_entry_lines jel
               JOIN journal_entries je ON je.aje_id = jel.aje_id
               WHERE je.job_id = ?
               ORDER BY jel.aje_id, jel.sort_order""",
            (job["job_id"],),
        )
    ]
    notes = [
        dict(r) for r in conn.execute(
            "SELECT * FROM notes WHERE job_id = ? ORDER BY created_at",
            (job["job_id"],),
        )
    ]
    account_groups = [
        dict(r) for r in conn.execute(
            "SELECT * FROM account_groups WHERE job_id = ? ORDER BY sort_order",
            (job["job_id"],),
        )
    ]
    account_group_members = []
    if account_groups:
        group_ids = [g["group_id"] for g in account_groups]
        placeholders = ",".join("?" * len(group_ids))
        account_group_members = [
            dict(r) for r in conn.execute(
                f"SELECT * FROM account_group_members WHERE group_id IN ({placeholders})",
                group_ids,
            )
        ]

    consolidation_members = [
        dict(r) for r in conn.execute(
            "SELECT * FROM consolidation_members WHERE job_id = ? ORDER BY sort_order",
            (job["job_id"],),
        )
    ] if job.get("entity_type") == "Consolidated" else []

    workpaper_lines = [
        dict(r) for r in conn.execute(
            "SELECT * FROM workpaper_lines WHERE job_id = ? ORDER BY workpaper, sort_order",
            (job["job_id"],),
        )
    ]

    consolidation_entries = [
        dict(r) for r in conn.execute(
            "SELECT * FROM consolidation_entries WHERE job_id = ? ORDER BY workpaper, entry_number",
            (job["job_id"],),
        )
    ]
    consolidation_entry_lines = []
    if consolidation_entries:
        entry_ids = [e["entry_id"] for e in consolidation_entries]
        placeholders = ",".join("?" * len(entry_ids))
        consolidation_entry_lines = [
            dict(r) for r in conn.execute(
                f"SELECT * FROM consolidation_entry_lines WHERE entry_id IN ({placeholders}) "
                f"ORDER BY entry_id, sort_order",
                entry_ids,
            )
        ]

    consolidation_account_groups = [
        dict(r) for r in conn.execute(
            "SELECT * FROM consolidation_account_groups WHERE job_id = ? ORDER BY sort_order",
            (job["job_id"],),
        )
    ]
    consolidation_group_members = []
    if consolidation_account_groups:
        group_ids = [g["group_id"] for g in consolidation_account_groups]
        placeholders = ",".join("?" * len(group_ids))
        consolidation_group_members = [
            dict(r) for r in conn.execute(
                f"SELECT * FROM consolidation_group_members WHERE group_id IN ({placeholders})",
                group_ids,
            )
        ]

    # Full activity log, including its hash chain, travels with the package —
    # the .atbr.xlsx is what students actually submit, so the audit trail has
    # to be verifiable from that file alone, not just the original .atbw.
    activity_log = [
        dict(r) for r in conn.execute(
            "SELECT * FROM activity_log WHERE job_id = ? ORDER BY rowid",
            (job["job_id"],),
        )
    ]

    payload = {
        "app":            "ATBWorkup",
        "schema_version": "2.0",
        "job_id":         job["job_id"],
        "job":            {k: job.get(k) for k in (
                              "client_name", "entity_name", "tax_year", "entity_type",
                              "prepared_by", "reviewer", "workpaper_folder",
                              "accounting_system", "status", "schema_version",
                              "app_version", "created_at", "updated_at",
                              "is_rollforward", "prior_year_job_id",
                              "finalized_at", "finalized_by",
                          )},
        "accounts":             raw_accounts,
        "tax_lines":            tax_lines,
        "mappings":             mappings,
        "entries":              entries,
        "entry_lines":          entry_lines,
        "notes":                notes,
        "activity_log":         activity_log,
        "account_groups":         account_groups,
        "account_group_members":  account_group_members,
        "consolidation_members":  consolidation_members,
        "workpaper_lines":        workpaper_lines,
        "consolidation_entries":       consolidation_entries,
        "consolidation_entry_lines":   consolidation_entry_lines,
        "consolidation_account_groups": consolidation_account_groups,
        "consolidation_group_members":  consolidation_group_members,
    }
    content = json.dumps(payload, default=str, sort_keys=True)
    # Excel's per-cell string limit is 32,767 chars.  Split into safe chunks
    # across consecutive rows; the importer concatenates them on read.
    _CHUNK = 32000
    for i, start in enumerate(range(0, max(len(content), 1), _CHUNK), start=1):
        ws.cell(row=i, column=1, value=content[start : start + _CHUNK])
    return content


# ── Hidden tab: __manifest ────────────────────────────────────────────────────

def _write_manifest_tab(wb, conn, job: dict, version: int, performed_by: str,
                        checksum: str, filename: str, now_str: str):
    from atbworkup.models.activity import verify_activity_chain

    tip_row = conn.execute(
        "SELECT row_hash FROM activity_log WHERE job_id = ? ORDER BY rowid DESC LIMIT 1",
        (job["job_id"],),
    ).fetchone()
    count_row = conn.execute(
        "SELECT COUNT(*) FROM activity_log WHERE job_id = ?", (job["job_id"],)
    ).fetchone()
    chain_intact, break_id = verify_activity_chain(conn, job["job_id"])

    ws = wb.create_sheet("__manifest")
    manifest = {
        "app":              "ATBWorkup",
        "schema_version":   "1.0",
        "package_type":     "review",
        "version_number":   version,
        "file_name":        filename,
        "job_id":           job["job_id"],
        "client_name":      job["client_name"],
        "tax_year":         job["tax_year"],
        "entity_type":      job["entity_type"],
        "exported_by":      performed_by,
        "exported_at":      now_str,
        "checksum_sha256":  checksum,
        "checksum_covers":  "__data",
        "activity_log_row_count":    count_row[0] if count_row else 0,
        "activity_log_tip_hash":     (tip_row[0] if tip_row else None) or "",
        "activity_log_chain_intact": chain_intact,
        "activity_log_break_id":     break_id or "",
    }
    for r, (k, v) in enumerate(manifest.items(), start=1):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=str(v))


# ── Summary helpers ───────────────────────────────────────────────────────────

def _write_summary_section_header(ws, label: str, row: int, ncols: int | None = None):
    if ncols is None:
        ncols = 1 + len(_AMT_HDRS)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c, value=label if c == 1 else "")
        cell.font = Font(bold=True, color=_WHITE, size=10)
        cell.fill = _fill(_NAVY)
    ws.row_dimensions[row].height = 16


def _write_summary_line_row(ws, row_data: list, row: int,
                            bold: bool = False, bg: str | None = None,
                            fg: str = _BLACK):
    fill = _fill(bg) if bg else None
    for c, val in enumerate(row_data, start=1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font   = Font(bold=bold, size=10, color=fg)
        cell.border = _thin_border
        if fill:
            cell.fill = fill
        if c > 1 and val is not None:
            cell.number_format = '#,##0.00;[Red](#,##0.00)'
            cell.alignment     = Alignment(horizontal="right")
    ws.row_dimensions[row].height = 15


# ── Shared helpers ────────────────────────────────────────────────────────────

def _write_sectioned_accounts(ws, accounts: list, headers: list,
                              start_row: int) -> tuple[int, str]:
    """
    Write accounts grouped by section → tax line, with subtotals per section.
    Returns (next_row, last_section_name).
    """
    r = start_row

    # Build ordered section → line → accounts structure preserving sort
    section_order: list[str] = []
    sections: dict[str, dict[str, list]] = {}
    for a in accounts:
        sec = a.get("section") or "Other"
        line = a.get("line_name") or "Unmapped"
        if sec not in sections:
            section_order.append(sec)
            sections[sec] = {}
        if line not in sections[sec]:
            sections[sec][line] = []
        sections[sec][line].append(a)

    last_section = ""
    for sec in section_order:
        line_groups = sections[sec]
        sec_accounts = [a for accts in line_groups.values() for a in accts]

        # Section header (navy)
        _write_section_row(ws, sec, len(headers), r, bg=_NAVY, fg=_WHITE, bold=True)
        r += 1

        # Lines within section
        for line_name, accts in line_groups.items():
            _write_section_row(ws, f"  {line_name}", len(headers), r,
                               bg=_LIGHT, fg=_NAVY, bold=False)
            r += 1
            for a in accts:
                row_data = [a.get("account_number") or "", f"    {a['account_name']}",
                            a.get("line_name") or ""]
                for key in _AMT_KEYS:
                    v = a[key]
                    row_data.append(v if v != 0 else None)
                _write_data_row(ws, row_data, r, amount_start_col=4)
                r += 1

        # Section subtotal
        totals = {k: round(sum(a[k] for a in sec_accounts), 2) for k in _AMT_KEYS}
        total_row = ["", f"Total {sec}", ""]
        for key in _AMT_KEYS:
            v = totals[key]
            total_row.append(v if v != 0 else None)
        _write_total_row(ws, total_row, r, amount_start_col=4)
        r += 2  # spacer after each section

        last_section = sec

    return r, last_section


def _set_tb_col_widths(ws):
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 28
    for i in range(4, 4 + len(_AMT_HDRS)):
        ws.column_dimensions[get_column_letter(i)].width = 14


_thin        = Side(style="thin", color="CCCCCC")
_thin_border = Border(bottom=_thin)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _write_header_row(ws, headers: list[str], row: int):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font      = Font(bold=True, color=_WHITE, size=10)
        cell.fill      = _fill(_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18


def _period_label(job: dict, as_of: bool) -> str:
    year = job.get("tax_year", "")
    return f"As of December 31, {year}" if as_of else f"For the Year Ended December 31, {year}"


def _write_report_title(ws, job: dict, statement_label: str, ncols: int, as_of: bool) -> int:
    """
    Classic 3-line statement header: client name / statement title / period
    (point-in-time "As of ..." for the Balance Sheet, "For the Year Ended
    ..." for period statements) — so a tab reads on its own, the way a
    printed financial statement does, instead of opening on a bare row of
    column labels. Returns the next free row.
    """
    rows = [
        (job.get("client_name", ""), Font(bold=True, color=_WHITE, size=13), _NAVY, 22),
        (statement_label,            Font(bold=True, color=_NAVY,  size=11), _PLAT, 18),
        (_period_label(job, as_of),  Font(italic=True, color=_NAVY, size=10), None, 16),
    ]
    for i, (text, font, bg, height) in enumerate(rows, start=1):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=ncols)
        cell = ws.cell(row=i, column=1, value=text)
        cell.font      = font
        if bg:
            cell.fill = _fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[i].height = height
    return 4


# ── Column-staggered subtotal layout (classified BS / multi-step IS) ─────────
# Textbook classified-statement style: individual accounts sit one column in
# from their group's subtotal, and each higher tier of subtotal steps one
# column further right — so the eye can trace "detail → group → grand total"
# without re-reading labels.
#   Col A  labels (indented per level)
#   Col B  individual account amounts
#   Col C  group/bucket subtotals (Total Current Assets, Total Fixed Assets, ...)
#   Col D  mid-level rollups (Total Noncurrent Assets, Total Liabilities,
#          Total Equity, Gross Profit)
#   Col E  grand totals (TOTAL ASSETS, TOTAL LIABILITIES & EQUITY, Net Income)
_COL_LABEL, _COL_ACCOUNT, _COL_GROUP, _COL_MID, _COL_GRAND = 1, 2, 3, 4, 5

_SUBTOTAL_BORDER = Border(bottom=Side(style="thin", color="000000"))
_GRAND_BORDER    = Border(top=Side(style="thin", color="000000"),
                          bottom=Side(style="double", color="000000"))


def _label_cell(ws, row: int, text: str, indent: int = 0, bold: bool = False):
    cell = ws.cell(row=row, column=_COL_LABEL, value=("  " * indent) + text)
    cell.font = Font(bold=bold, size=10)
    return cell


def _amount_cell(ws, row: int, col: int, value: float, bold: bool = False,
                 border: Border | None = None, always_show: bool = False):
    show = round(value, 2) if (value or always_show) else None
    cell = ws.cell(row=row, column=col, value=show)
    cell.font = Font(bold=bold, size=10)
    if border is not None:
        cell.border = border
    if show is not None:
        cell.number_format = '#,##0.00;[Red](#,##0.00)'
        cell.alignment     = Alignment(horizontal="right")
    return cell


def _write_group_accounts(ws, accounts: list, start_row: int, sign: float = 1.0,
                          ncols: int = 5) -> tuple[int, float, set]:
    """
    Write one GAAP bucket: section banner(s) → account rows (col B) → one
    subtotal per section (col C). Returns (next_row, bucket_total,
    section_names) — callers use section_names to decide whether a further
    mid-level rollup row would just duplicate the one section subtotal
    already written (skip it when there's only one section).
    """
    sections: dict[str, list] = {}
    order: list[str] = []
    for a in accounts:
        sec = a.get("section") or "Other"
        if sec not in sections:
            order.append(sec)
            sections[sec] = []
        sections[sec].append(a)

    r = start_row
    bucket_total = 0.0
    for sec in order:
        accts = sections[sec]
        _write_summary_section_header(ws, sec, r, ncols=ncols)
        r += 1
        for a in accts:
            _label_cell(ws, r, a["account_name"], indent=1)
            _amount_cell(ws, r, _COL_ACCOUNT, sign * a.get("final", 0.0))
            r += 1
        sec_total = sign * sum(a.get("final", 0.0) for a in accts)
        _label_cell(ws, r, f"Total {sec}", bold=True)
        _amount_cell(ws, r, _COL_GROUP, sec_total, bold=True, border=_SUBTOTAL_BORDER)
        r += 2
        bucket_total += sec_total

    return r, round(bucket_total, 2), set(order)


def _write_section_row(ws, label: str, ncols: int, row: int,
                       bg: str, fg: str, bold: bool):
    fill = _fill(bg)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c, value=label if c == 2 else "")
        cell.font = Font(bold=bold, color=fg, size=10)
        cell.fill = fill
    ws.row_dimensions[row].height = 16


def _write_data_row(ws, row_data: list, row: int,
                    amount_start_col: int, amount_cols: int | None = None):
    if amount_cols is None:
        amount_cols = len(row_data) - (amount_start_col - 1)
    for c, val in enumerate(row_data, start=1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font   = Font(size=10)
        cell.border = _thin_border
        if c >= amount_start_col and val is not None:
            cell.number_format = '#,##0.00;[Red](#,##0.00)'
            cell.alignment     = Alignment(horizontal="right")
    ws.row_dimensions[row].height = 15


def _write_total_row(ws, row_data: list, row: int, amount_start_col: int):
    fill = _fill(_PLAT)
    for c, val in enumerate(row_data, start=1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = Font(bold=True, size=10)
        cell.fill = fill
        if c >= amount_start_col and val is not None:
            cell.number_format = '#,##0.00;[Red](#,##0.00)'
            cell.alignment     = Alignment(horizontal="right")
    ws.row_dimensions[row].height = 16


def _sha256(content: str) -> str:
    return hashlib.sha256(content.encode("utf-8")).hexdigest()


def _now() -> str:
    return datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
