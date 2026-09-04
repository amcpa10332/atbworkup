"""
Trial balance Excel parser — no Qt dependency.

Responsibilities:
  - Load a workbook and expose sheet names
  - Read raw rows from a selected sheet
  - Auto-detect the header row
  - Parse rows into ParsedAccount records using caller-supplied column mapping
  - Handle: single-column and two-column (Dr/Cr) layouts
  - Handle: parenthetical negatives  (1,234.56)  →  -1234.56
  - Skip blank / subtotal rows
  - Compute running debit/credit totals for the preview
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class ParsedAccount:
    source_row: int        # 1-based row number in the sheet
    account_number: str    # may be empty string
    account_name: str
    pbc_balance: float     # DR = positive, CR = negative


@dataclass
class ParseResult:
    accounts: list[ParsedAccount] = field(default_factory=list)
    total_debits: float = 0.0
    total_credits: float = 0.0
    skipped_rows: int = 0

    @property
    def is_balanced(self) -> bool:
        return abs(self.total_debits + self.total_credits) < 0.005


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def get_sheet_names(path: str | Path) -> list[str]:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def read_raw_rows(path: str | Path, sheet_name: str) -> list[list[Any]]:
    """
    Return all rows as a list-of-lists of cell values.
    Trailing empty cells are kept so column indices stay stable.
    """
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    wb.close()
    return rows


_HEADER_KEYWORDS = {
    "debit", "credit", "balance", "amount", "dr", "cr", "net",
    "debits", "credits", "acct", "account", "description", "name",
}


def detect_header_row(rows: list[list[Any]], name_col: int) -> int:
    """
    Return 0-based index of the most likely header row.

    Strategy (in priority order):
    1. Find the first row (within first 30) that contains a cell matching
       a known column-header keyword ("Debit", "Credit", "Balance", etc.).
       This handles TBs with title rows above the actual header.
    2. Fall back to first row where name_col is a non-empty string
       that is not a number.
    """
    for i, row in enumerate(rows[:30]):
        for cell in row:
            if isinstance(cell, str) and cell.strip().lower() in _HEADER_KEYWORDS:
                return i

    for i, row in enumerate(rows[:30]):
        val = _cell_str(row, name_col)
        if val and not _try_parse_amount(val)[1]:
            return i

    return 0


def parse_accounts(
    rows: list[list[Any]],
    *,
    header_row: int,
    name_col: int,
    balance_col: int | None = None,
    debit_col: int | None = None,
    credit_col: int | None = None,
    number_col: int | None = None,
) -> ParseResult:
    """
    Parse rows into ParsedAccount records.

    Layout modes:
      Single-column:  supply balance_col.  Positive = DR, negative = CR.
      Two-column:     supply debit_col and credit_col.
                      stored amount = debit_amount - credit_amount  (DR+, CR-)

    Rows are skipped when:
      - name column is blank
      - row index <= header_row
      - all balance cells are blank
    """
    result = ParseResult()

    for row_idx, row in enumerate(rows):
        if row_idx <= header_row:
            continue

        name = _cell_str(row, name_col).strip()
        if not name:
            result.skipped_rows += 1
            continue

        number = _cell_str(row, number_col).strip() if number_col is not None else ""

        # --- parse amount ---
        if balance_col is not None:
            raw = _cell_val(row, balance_col)
            if raw is None or str(raw).strip() == "":
                result.skipped_rows += 1
                continue
            amount, ok = _try_parse_amount(str(raw))
            if not ok:
                result.skipped_rows += 1
                continue
        elif debit_col is not None and credit_col is not None:
            dr_raw = _cell_val(row, debit_col)
            cr_raw = _cell_val(row, credit_col)
            if dr_raw is None and cr_raw is None:
                result.skipped_rows += 1
                continue
            dr, _ = _try_parse_amount(str(dr_raw or "0"))
            cr, _ = _try_parse_amount(str(cr_raw or "0"))
            amount = dr - cr
        else:
            raise ValueError("Must supply either balance_col or both debit_col and credit_col.")

        acct = ParsedAccount(
            source_row=row_idx + 1,  # 1-based
            account_number=number,
            account_name=name,
            pbc_balance=round(amount, 2),
        )
        result.accounts.append(acct)

        if amount >= 0:
            result.total_debits += amount
        else:
            result.total_credits += amount

    result.total_debits = round(result.total_debits, 2)
    result.total_credits = round(result.total_credits, 2)
    return result


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

_PAREN_RE = re.compile(r"^\(([0-9,. ]+)\)$")


def _try_parse_amount(s: str) -> tuple[float, bool]:
    """
    Parse a string to float.  Returns (value, success).
    Handles:  1234.56  |  1,234.56  |  (1,234.56)  |  -1234.56
    """
    s = s.strip().replace(" ", "")
    if not s:
        return 0.0, False
    # parenthetical negative
    m = _PAREN_RE.match(s)
    if m:
        try:
            return -float(m.group(1).replace(",", "")), True
        except ValueError:
            return 0.0, False
    # strip commas
    s = s.replace(",", "")
    try:
        return float(s), True
    except ValueError:
        return 0.0, False


def _cell_val(row: list[Any], col: int) -> Any:
    if col < 0 or col >= len(row):
        return None
    return row[col]


def _cell_str(row: list[Any], col: int) -> str:
    v = _cell_val(row, col)
    return "" if v is None else str(v)
