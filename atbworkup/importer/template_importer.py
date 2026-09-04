"""
Tax line template Excel importer.

Expected workbook layout
------------------------
Sheet "Info":
    Row 2: ("Template Name", <name>)
    Row 3: ("Entity Type Code", <code>)   — optional; falls back to template name
    Row 4: ("Description", <desc>)

Content sheets (one per financial statement):
    Sheet name  = financial_statement value (e.g. "Balance Sheet", "Profit & Loss")
    Row 1       = header: section_sort_order | section | line_code | line_name | sort_order
    Row 2+      = data rows; blank rows are silently skipped
"""
from __future__ import annotations

import uuid
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_template_excel(xlsx_path: str | Path) -> dict:
    """
    Parse a tax line template Excel file.

    Returns
    -------
    {
        "template_name": str,
        "entity_type_code": str,   # same as template_name when not specified
        "description": str,
        "lines": [
            {
                "financial_statement": str,
                "section": str,
                "section_sort_order": int,
                "line_code": str,
                "line_name": str,
                "sort_order": int,
                "is_active": int,   # always 1
            },
            ...
        ]
    }

    Raises
    ------
    ValueError  with a descriptive message on any validation error.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise ValueError(f"File not found: {xlsx_path}")

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot open workbook: {exc}") from exc

    # ── Info sheet ──────────────────────────────────────────────────────────
    if "Info" not in wb.sheetnames:
        raise ValueError("Missing 'Info' sheet")

    info_ws = wb["Info"]
    info_rows = list(info_ws.iter_rows(min_row=2, max_row=4, values_only=True))

    def _cell(row_tuple, col_index: int) -> str:
        if row_tuple is None:
            return ""
        val = row_tuple[col_index] if len(row_tuple) > col_index else None
        return str(val).strip() if val is not None else ""

    template_name   = _cell(info_rows[0] if len(info_rows) > 0 else None, 1)
    entity_type_raw = _cell(info_rows[1] if len(info_rows) > 1 else None, 1)
    description     = _cell(info_rows[2] if len(info_rows) > 2 else None, 1)

    if not template_name:
        raise ValueError("'Info' sheet row 2 must contain a non-empty Template Name")

    entity_type_code = entity_type_raw if entity_type_raw else template_name

    # ── Content sheets ──────────────────────────────────────────────────────
    content_sheet_names = [n for n in wb.sheetnames if n != "Info"]
    if not content_sheet_names:
        raise ValueError("No financial statement sheets found")

    lines: list[dict] = []

    for sheet_name in content_sheet_names:
        ws = wb[sheet_name]
        financial_statement = sheet_name

        rows_iter = ws.iter_rows(min_row=2, values_only=True)  # skip header row 1
        sheet_lines: list[dict] = []

        for row in rows_iter:
            # Skip completely blank rows
            if not any(cell is not None and str(cell).strip() != "" for cell in row):
                continue

            # Expected columns: section_sort_order | section | line_code | line_name | sort_order
            def _v(idx: int) -> str:
                val = row[idx] if len(row) > idx else None
                return str(val).strip() if val is not None else ""

            section_sort_raw = _v(0)
            section          = _v(1)
            line_code        = _v(2)
            line_name        = _v(3)
            sort_raw         = _v(4)

            # Validate integer fields — skip malformed rows silently
            try:
                section_sort_order = int(float(section_sort_raw))
            except (ValueError, TypeError):
                continue
            try:
                sort_order = int(float(sort_raw))
            except (ValueError, TypeError):
                continue

            # Validate required string fields
            if not line_code or not line_name:
                continue

            sheet_lines.append({
                "financial_statement": financial_statement,
                "section":             section,
                "section_sort_order":  section_sort_order,
                "line_code":           line_code,
                "line_name":           line_name,
                "sort_order":          sort_order,
                "is_active":           1,
            })

        if not sheet_lines:
            raise ValueError(
                f"Sheet '{sheet_name}' has no valid data rows. "
                "Each row needs: section_sort_order (int), section, line_code, line_name, sort_order (int)."
            )

        lines.extend(sheet_lines)

    wb.close()

    return {
        "template_name":    template_name,
        "entity_type_code": entity_type_code,
        "description":      description,
        "lines":            lines,
    }


def import_template_from_excel(xlsx_path: str | Path) -> dict:
    """
    Parse an Excel template file and write it to the settings database.

    Returns the parsed dict (same shape as parse_template_excel).

    Raises
    ------
    ValueError  on bad file content or if the entity_type_code already exists
                as a non-built-in (custom) template.
    """
    parsed = parse_template_excel(xlsx_path)
    _upsert_template(parsed)
    return parsed


# ---------------------------------------------------------------------------
# Internal DB write
# ---------------------------------------------------------------------------

def _upsert_template(parsed: dict) -> None:
    """Write parsed template lines to the settings DB."""
    from atbworkup.db.settings import settings_connection

    entity_type_code = parsed["entity_type_code"]
    template_name    = parsed["template_name"]
    description      = parsed.get("description", "")
    lines            = parsed["lines"]

    with settings_connection() as conn:
        # Detect whether template_name / is_builtin columns exist
        tlt_cols = {row[1] for row in conn.execute("PRAGMA table_info(tax_line_templates)")}
        has_template_name = "template_name" in tlt_cols
        has_is_builtin    = "is_builtin" in tlt_cols

        # Guard: do not overwrite a built-in template
        if has_is_builtin:
            existing = conn.execute(
                "SELECT is_builtin FROM tax_line_templates WHERE entity_type = ? LIMIT 1",
                (entity_type_code,),
            ).fetchone()
            if existing is not None and existing[0]:
                raise ValueError(
                    f"Entity type code '{entity_type_code}' belongs to a built-in template "
                    "and cannot be overwritten via import."
                )

        # Delete any existing rows for this entity type
        conn.execute(
            "DELETE FROM tax_line_templates WHERE entity_type = ?",
            (entity_type_code,),
        )

        # Build INSERT
        if has_template_name and has_is_builtin:
            sql = (
                "INSERT INTO tax_line_templates "
                "(template_id, entity_type, template_name, financial_statement, "
                " section, section_sort_order, line_code, line_name, sort_order, is_active, is_builtin) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?)"
            )
            rows = [
                (
                    uuid.uuid4().hex,
                    entity_type_code,
                    template_name,
                    ln["financial_statement"],
                    ln["section"],
                    ln["section_sort_order"],
                    ln["line_code"],
                    ln["line_name"],
                    ln["sort_order"],
                    ln["is_active"],
                    0,  # is_builtin = False for imported templates
                )
                for ln in lines
            ]
        elif has_template_name:
            sql = (
                "INSERT INTO tax_line_templates "
                "(template_id, entity_type, template_name, financial_statement, "
                " section, section_sort_order, line_code, line_name, sort_order, is_active) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)"
            )
            rows = [
                (
                    uuid.uuid4().hex,
                    entity_type_code,
                    template_name,
                    ln["financial_statement"],
                    ln["section"],
                    ln["section_sort_order"],
                    ln["line_code"],
                    ln["line_name"],
                    ln["sort_order"],
                    ln["is_active"],
                )
                for ln in lines
            ]
        else:
            sql = (
                "INSERT INTO tax_line_templates "
                "(template_id, entity_type, financial_statement, "
                " section, section_sort_order, line_code, line_name, sort_order, is_active) "
                "VALUES (?,?,?,?,?,?,?,?,?)"
            )
            rows = [
                (
                    uuid.uuid4().hex,
                    entity_type_code,
                    ln["financial_statement"],
                    ln["section"],
                    ln["section_sort_order"],
                    ln["line_code"],
                    ln["line_name"],
                    ln["sort_order"],
                    ln["is_active"],
                )
                for ln in lines
            ]

        conn.executemany(sql, rows)
