"""
Open an .atbr.xlsx review package as a working binder:
  1. Read __manifest and verify the SHA-256 checksum.
  2. Parse the __data JSON snapshot (must be schema_version "2.0").
  3. Hydrate a fresh ephemeral .atbw in %TEMP%\\ATBWorkup\\.
  4. Return (temp_atbw_path, job_dict).

The temp .atbw is an ephemeral cache — the .atbr.xlsx is the source of truth.
"""
from __future__ import annotations

import hashlib
import json
from pathlib import Path


def open_from_package(xlsx_path: str | Path, performed_by: str) -> tuple[Path, dict]:
    """
    Open a .atbr.xlsx package.  Returns (temp_atbw_path, job_dict).

    Raises ValueError if:
      - __manifest or __data sheets are missing
      - the SHA-256 checksum does not match (file may have been externally edited)
      - schema_version != "2.0" (old single-session format, not re-importable)
    """
    import openpyxl

    xlsx_path = Path(xlsx_path)
    # Do NOT use read_only=True — ReadOnlyWorksheet.cell() access is unreliable
    # and can silently return None for long string cells, breaking checksum verification.
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    if "__manifest" not in wb.sheetnames:
        raise ValueError(
            "Not a valid ATBWorkup package: __manifest sheet is missing.\n"
            "This file may not have been created by ATBWorkup."
        )
    if "__data" not in wb.sheetnames:
        raise ValueError(
            "Not a valid ATBWorkup package: __data sheet is missing.\n"
            "This file may not have been created by ATBWorkup."
        )

    manifest: dict[str, str] = {}
    for row in wb["__manifest"].iter_rows(values_only=True):
        if row and row[0]:
            manifest[str(row[0])] = str(row[1]) if row[1] is not None else ""

    # __data is chunked across consecutive rows (32 000 chars each) to stay
    # under Excel's 32 767 per-cell string limit.  Concatenate all chunks.
    data_content: str = ""
    for row in wb["__data"].iter_rows(values_only=True):
        if row and row[0] is not None:
            data_content += str(row[0])

    # --- checksum verification ---
    actual   = hashlib.sha256(data_content.encode("utf-8")).hexdigest()
    expected = manifest.get("checksum_sha256", "")
    if actual != expected:
        # Diagnostic: show first 200 chars of what was actually read
        preview = repr(data_content[:200]) if data_content else "<empty>"
        raise ValueError(
            "Package integrity check failed — the file may have been modified "
            "outside of ATBWorkup.\n\n"
            f"Expected checksum: {expected}\n"
            f"Actual checksum:   {actual}\n"
            f"Data length read:  {len(data_content)} chars\n"
            f"Data preview:      {preview}\n\n"
            "If you believe this file is correct, contact the person who sent it."
        )

    data = json.loads(data_content)

    schema_version = data.get("schema_version", "1.0")
    if schema_version != "2.0":
        raise ValueError(
            f"This package uses an older format (schema_version={schema_version!r}) "
            "and cannot be reopened as a working file.  "
            "Please ask the preparer to re-export it from the current version of ATBWorkup."
        )

    job_id    = data["job_id"]
    job_meta  = data["job"]

    from atbworkup.utils.naming import temp_atbw_path
    from atbworkup.models.job import create_workup, get_job
    from atbworkup.db.connection import db_connection
    from atbworkup.importer.hydration import hydrate_binder
    from atbworkup.models.activity import log_activity

    temp_path = temp_atbw_path(job_id)

    # Remove any stale temp file from a previous session
    if temp_path.exists():
        temp_path.unlink()

    create_workup(temp_path, job_meta, job_id=job_id)

    with db_connection(temp_path) as conn:
        hydrate_binder(conn, data)
        log_activity(
            conn,
            job_id=job_id,
            event_type="opened_from_package",
            description=f"Opened from {xlsx_path.name}",
            performed_by=performed_by,
        )

    job = get_job(temp_path)
    return temp_path, job
