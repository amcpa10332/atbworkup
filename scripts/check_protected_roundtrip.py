"""
Diagnostic: test round-trip WITH veryHidden + sheet protection + workbook lock.
Run: python scripts/check_protected_roundtrip.py
"""
import hashlib
import json
import sys
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.workbook.protection import WorkbookProtection

_SHEET_LOCK_PASSWORD    = "atbwi-sl-1"
_WORKBOOK_LOCK_PASSWORD = "atbwi-wb-1"

payload = {
    "app": "ATBWorkup",
    "schema_version": "2.0",
    "job_id": "abc123",
    "job": {
        "client_name": "Test & Co",   # & is an XML-special char
        "tax_year": 2024,
        "workpaper_folder": r"C:\Users\AustinMalone\OneDrive - zbcpa.tax\Desktop",
        "reviewer": None,
        "is_rollforward": 0,
    },
    "accounts": [{"account_id": "x", "pbc_balance": 5000.0, "is_mapped": 0,
                  "account_name": "Cash & Equivalents"}],
    "tax_lines": [],
    "mappings": [],
    "entries": [],
    "entry_lines": [],
    "notes": [{"note_id": "n1", "body": "Line 1\nLine 2", "status": "Open"}],
}

content = json.dumps(payload, default=str, sort_keys=True)
expected = hashlib.sha256(content.encode("utf-8")).hexdigest()
print(f"Content length : {len(content)}")
print(f"Written hash   : {expected}")
print(f"First 120 chars: {content[:120]}")

p = Path(tempfile.gettempdir()) / "atbtest_protected.xlsx"
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Add a visible sheet (like the real exporter always does)
ws_cover = wb.create_sheet("Cover")
ws_cover.cell(1, 1, value="Cover Page")

ws_data = wb.create_sheet("__data")
ws_data.cell(1, 1, value=content)

ws_manifest = wb.create_sheet("__manifest")
ws_manifest.cell(1, 1, value="checksum_sha256")
ws_manifest.cell(1, 2, value=expected)

for ws in wb.worksheets:
    if ws.title.startswith("__"):
        ws.sheet_state = "veryHidden"
        ws.protection.sheet               = True
        ws.protection.password            = _SHEET_LOCK_PASSWORD
        ws.protection.selectLockedCells   = False
        ws.protection.selectUnlockedCells = False

wb.security = WorkbookProtection(
    lockStructure=True,
    workbookPassword=_WORKBOOK_LOCK_PASSWORD,
)
wb.save(str(p))
print(f"Saved to {p}")

# --- read back ---
wb2 = openpyxl.load_workbook(str(p), data_only=True)
print(f"Sheets visible in load: {wb2.sheetnames}")

read_content = ""
for row in wb2["__data"].iter_rows(min_row=1, max_row=1, values_only=True):
    if row and row[0] is not None:
        read_content = str(row[0])

manifest = {}
for row in wb2["__manifest"].iter_rows(values_only=True):
    if row and row[0]:
        manifest[str(row[0])] = str(row[1]) if row[1] is not None else ""

stored_checksum = manifest.get("checksum_sha256", "")
actual = hashlib.sha256(read_content.encode("utf-8")).hexdigest()
print(f"Stored hash    : {stored_checksum}")
print(f"Computed hash  : {actual}")
print(f"Expected hash  : {expected}")
print(f"Hash match     : {expected == actual}")
print(f"Content match  : {content == read_content}")
if content != read_content:
    print(f"  written len={len(content)} read len={len(read_content)}")
    print(f"  Read first 120: {read_content[:120]}")

p.unlink()
sys.exit(0 if expected == actual else 1)
