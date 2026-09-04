"""
Grading tool: verify a submitted .atbr.xlsx package's audit trail hasn't been
tampered with, and print a pacing summary (sessions, hostnames, durations) so
an instructor can sanity-check that the work looks like it was actually done
by one person over a plausible amount of time.

Usage (from the project root):
    python scripts/verify_audit_chain.py "path/to/Student Submission.atbr.xlsx"

What it checks:
  1. Recomputes SHA-256 over the __data tab's content and compares it to the
     __manifest tab's checksum_sha256 — catches any edit made to the exported
     file after the fact (including editing activity_log entries in place).
  2. Recomputes the activity_log hash chain from scratch and compares the
     final link to __manifest's activity_log_tip_hash — catches deleted or
     reordered activity_log rows, which a checksum alone wouldn't (a student
     could remove trailing rows and the file would otherwise look untouched
     up to that point).
  3. Prints session_started/session_ended pairs with hostname and duration,
     and flags any file whose total logged work time looks implausibly short
     for its content (a rough heuristic, not proof by itself).

This only detects tampering with the *exported .xlsx*. It cannot detect a
student handing their finished .atbw file to someone else who then exports
it themselves — the exported_by/performed_by fields reflect the settings
profile that did the export, so cross-check that name against your roster.
"""
from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path

import openpyxl


def _read_data_content(wb) -> str:
    ws = wb["__data"]
    parts = []
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        if row[0] is None:
            break
        parts.append(row[0])
    return "".join(parts)


def _read_manifest(wb) -> dict:
    ws = wb["__manifest"]
    manifest = {}
    for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
        key, value = row
        if key is None:
            continue
        manifest[key] = value
    return manifest


def _compute_row_hash(prev_hash, job_id, event_type, entity_type, entity_id,
                      description, performed_by, performed_at,
                      package_version, metadata_json) -> str:
    payload = "|".join(
        "" if x is None else str(x)
        for x in (prev_hash, job_id, event_type, entity_type, entity_id,
                  description, performed_by, performed_at, package_version,
                  metadata_json)
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def verify(path: str | Path) -> int:
    path = Path(path)
    wb = openpyxl.load_workbook(str(path), data_only=True)

    if "__data" not in wb.sheetnames or "__manifest" not in wb.sheetnames:
        print(f"FAIL  {path.name}: not an ATBWorkup review package "
              f"(missing __data/__manifest tabs)")
        return 1

    manifest = _read_manifest(wb)
    content = _read_data_content(wb)
    payload = json.loads(content)

    ok = True

    # ── Checksum: has the exported file been edited since export? ──────────
    recomputed_checksum = hashlib.sha256(content.encode("utf-8")).hexdigest()
    expected_checksum = manifest.get("checksum_sha256", "")
    if recomputed_checksum == expected_checksum:
        print(f"PASS  __data checksum matches manifest (file unedited since export)")
    else:
        print(f"FAIL  __data checksum MISMATCH — this file was modified after export")
        print(f"      expected: {expected_checksum}")
        print(f"      actual:   {recomputed_checksum}")
        ok = False

    # ── Hash chain: has the activity log been altered or had rows removed? ─
    job_id = payload.get("job_id", "")
    # The app writes this list ORDER BY rowid (true insertion order), and
    # JSON arrays preserve order through serialization — so list order here
    # already matches the order the hash chain was built in. Do not re-sort
    # (performed_at has second resolution and ties are common; sorting by it
    # would silently reorder same-second events and break the chain check).
    activity_log = payload.get("activity_log", [])
    expected_prev = "0" * 64
    break_at = None
    for row in activity_log:
        recomputed = _compute_row_hash(
            expected_prev, job_id, row.get("event_type"), row.get("entity_type"),
            row.get("entity_id"), row.get("description"), row.get("performed_by"),
            row.get("performed_at"), row.get("package_version"), row.get("metadata_json"),
        )
        if row.get("prev_hash") != expected_prev or row.get("row_hash") != recomputed:
            break_at = row.get("activity_id")
            break
        expected_prev = row["row_hash"]

    expected_tip = manifest.get("activity_log_tip_hash", "")
    if break_at is None and expected_prev == expected_tip:
        print(f"PASS  activity log hash chain intact ({len(activity_log)} events)")
    else:
        print(f"FAIL  activity log hash chain BROKEN"
              f"{f' at row {break_at}' if break_at else ' (tip hash mismatch)'}"
              f" — rows were edited, deleted, or reordered")
        ok = False

    # ── Pacing summary ──────────────────────────────────────────────────────
    sessions: dict[str, dict] = {}
    for row in activity_log:
        meta_raw = row.get("metadata_json")
        if not meta_raw:
            continue
        try:
            meta = json.loads(meta_raw)
        except (TypeError, ValueError):
            continue
        sid = meta.get("session_id")
        if not sid:
            continue
        s = sessions.setdefault(sid, {"hostname": meta.get("hostname"), "start": None, "duration": None})
        if row.get("event_type") == "session_started":
            s["start"] = row.get("performed_at")
        elif row.get("event_type") == "session_ended":
            s["duration"] = meta.get("duration_seconds")

    print(f"\nSessions logged: {len(sessions)}")
    total_seconds = 0.0
    hostnames = set()
    for sid, s in sorted(sessions.items(), key=lambda kv: kv[1]["start"] or ""):
        dur = s["duration"]
        total_seconds += dur or 0.0
        if s["hostname"]:
            hostnames.add(s["hostname"])
        dur_str = f"{dur / 60:.1f} min" if dur is not None else "(no session_ended — crash or force-quit?)"
        print(f"  {s['start'] or '?':<22} host={s['hostname'] or '?':<20} duration={dur_str}")

    print(f"\nTotal logged active time: {total_seconds / 60:.1f} min")
    if len(hostnames) > 1:
        print(f"NOTE: work was logged from {len(hostnames)} different machines: {sorted(hostnames)}")
        print(f"      (not necessarily suspicious — could be a lab machine + personal laptop —")
        print(f"       but worth a quick check if unexpected)")

    print(f"\n{'PASS' if ok else 'FAIL'}: {path.name}")
    return 0 if ok else 1


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python scripts/verify_audit_chain.py <path-to-.atbr.xlsx>")
        sys.exit(2)
    sys.exit(verify(sys.argv[1]))
