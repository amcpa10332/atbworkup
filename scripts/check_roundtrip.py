"""
Diagnostic: verify that openpyxl round-trips the __data cell content intact.
Run from the project root: python scripts/check_roundtrip.py
"""
import hashlib
import json
import sys
import tempfile
from pathlib import Path

import openpyxl

payload = {
    "app": "ATBWorkup",
    "schema_version": "2.0",
    "job_id": "abc123",
    "job": {
        "client_name": "Test Co",
        "tax_year": 2024,
        "workpaper_folder": r"C:\Users\AustinMalone\OneDrive - zbcpa.tax\Desktop",
    },
    "accounts": [{"account_id": "x", "pbc_balance": 5000.0, "is_mapped": 0}],
    "tax_lines": [],
    "mappings": [],
    "entries": [],
    "entry_lines": [],
    "notes": [],
}

content = json.dumps(payload, default=str, sort_keys=True)
expected = hashlib.sha256(content.encode("utf-8")).hexdigest()
print(f"Content length : {len(content)}")
print(f"Written hash   : {expected}")

p = Path(tempfile.gettempdir()) / "atbtest_roundtrip.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "__data"
ws.cell(1, 1, value=content)
wb.save(str(p))

wb2 = openpyxl.load_workbook(str(p), data_only=True)
read_content = ""
for row in wb2["__data"].iter_rows(min_row=1, max_row=1, values_only=True):
    if row and row[0] is not None:
        read_content = str(row[0])

actual = hashlib.sha256(read_content.encode("utf-8")).hexdigest()
print(f"Read hash      : {actual}")
print(f"Hash match     : {expected == actual}")
print(f"Content match  : {content == read_content}")

if content != read_content:
    print(f"\nLength diff: written={len(content)} read={len(read_content)}")
    for i, (a, b) in enumerate(zip(content, read_content)):
        if a != b:
            print(f"First diff at index {i}: written={repr(a)} read={repr(b)}")
            print(f"Context: ...{repr(content[max(0, i-30):i+30])}...")
            break

p.unlink()
sys.exit(0 if expected == actual else 1)
