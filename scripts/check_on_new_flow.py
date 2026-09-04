"""
Diagnostic: simulate the exact _on_new + closeEvent flow, then verify open_from_package.
Run: python scripts/check_on_new_flow.py
"""
import sys, hashlib, uuid, tempfile, json
sys.path.insert(0, '.')
from pathlib import Path

from atbworkup.db.settings import ensure_settings_db, set_settings_path
from atbworkup.db.connection import db_connection
from atbworkup.models.job import create_workup, get_job
from atbworkup.exporter.review_package import save_workup
from atbworkup.importer.package import open_from_package
from atbworkup.utils.naming import temp_atbw_path, suggested_filename

import openpyxl

tmp = Path(tempfile.mkdtemp())
set_settings_path(tmp / "settings.db")
ensure_settings_db()

meta = {
    "client_name": "Austin & Partners",
    "entity_name": "Austin Partners LLC",
    "tax_year": 2024,
    "entity_type": "1120S",
    "prepared_by": "Austin Malone",
    "reviewer": None,
    "workpaper_folder": str(tmp),
    "accounting_system": "QuickBooks",
}

# --- Simulate _on_new ---
job_id    = uuid.uuid4().hex
temp_path = temp_atbw_path(job_id)
xlsx_path = tmp / suggested_filename(meta["tax_year"], meta["client_name"])

print(f"1. create_workup -> {temp_path.name}")
create_workup(temp_path, meta, job_id=job_id)
job = get_job(temp_path)

print(f"2. save_workup (initial write) -> {xlsx_path.name}")
with db_connection(temp_path) as conn:
    save_workup(conn, job=job, output_path=xlsx_path, performed_by=meta["prepared_by"])

# Read back to confirm initial write
wb1 = openpyxl.load_workbook(str(xlsx_path), data_only=True)
dc1 = ""
for row in wb1["__data"].iter_rows(min_row=1, max_row=1, values_only=True):
    if row and row[0]: dc1 = str(row[0])
mf1 = {str(r[0]): str(r[1]) for r in wb1["__manifest"].iter_rows(values_only=True) if r and r[0]}
h1_stored  = mf1.get("checksum_sha256", "")
h1_actual  = hashlib.sha256(dc1.encode()).hexdigest()
print(f"   Initial write - stored={h1_stored[:16]}... actual={h1_actual[:16]}... match={h1_stored==h1_actual}")
print(f"   Data length: {len(dc1)} chars")

# --- Simulate closeEvent (auto-save) ---
print(f"3. _on_save (closeEvent auto-save) -> {xlsx_path.name}")
# This uses self._job which is `job` from get_job above
with db_connection(temp_path) as conn:
    save_workup(conn, job=job, output_path=xlsx_path, performed_by=meta["prepared_by"])

# Read back after auto-save
wb2 = openpyxl.load_workbook(str(xlsx_path), data_only=True)
dc2 = ""
for row in wb2["__data"].iter_rows(min_row=1, max_row=1, values_only=True):
    if row and row[0]: dc2 = str(row[0])
mf2 = {str(r[0]): str(r[1]) for r in wb2["__manifest"].iter_rows(values_only=True) if r and r[0]}
h2_stored  = mf2.get("checksum_sha256", "")
h2_actual  = hashlib.sha256(dc2.encode()).hexdigest()
print(f"   Auto-save    - stored={h2_stored[:16]}... actual={h2_actual[:16]}... match={h2_stored==h2_actual}")
print(f"   Data length: {len(dc2)} chars")
print(f"   Data same as initial: {dc1 == dc2}")

# Delete temp (closeEvent does this)
temp_path.unlink(missing_ok=True)

# --- Simulate open_from_package ---
print(f"4. open_from_package -> {xlsx_path.name}")
try:
    new_temp, reopened_job = open_from_package(xlsx_path, performed_by="Tester")
    print(f"   SUCCESS - client={reopened_job['client_name']}")
    new_temp.unlink(missing_ok=True)
except ValueError as e:
    print(f"   FAILED:\n{e}")
    sys.exit(1)

import shutil
shutil.rmtree(str(tmp))
print("\nAll OK.")
