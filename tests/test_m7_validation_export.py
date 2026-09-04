"""M7 tests: validation engine and review package export."""
import pytest
import uuid
import datetime
import json
import hashlib
from pathlib import Path

import openpyxl

from atbworkup.db.connection import db_connection
from atbworkup.models.job import get_job
from atbworkup.models.validation import run_all, all_pass
from atbworkup.exporter.review_package import (
    export_review_package, suggested_filename, next_version,
)
from atbworkup.db.settings import set_settings_path, ensure_settings_db
from atbworkup.models.mappings import get_tax_line_templates, upsert_tax_line, map_accounts


# ---------------------------------------------------------------------------
# Fixtures / helpers
# ---------------------------------------------------------------------------

def _now():
    return datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _insert_account(conn, job_id, *, name, balance=0.0, number=""):
    aid = uuid.uuid4().hex
    now = _now()
    conn.execute(
        """INSERT INTO accounts
               (account_id, job_id, account_number, account_name, account_type,
                pbc_balance, normal_balance, sort_order, is_mapped, created_at, updated_at)
           VALUES (?,?,?,?,?,?,?,?,0,?,?)""",
        (aid, job_id, number, name, "Asset", balance, "Debit", 0, now, now),
    )
    return aid


def _balanced_tb(conn, job_id):
    """Insert two accounts with offsetting balances and map them."""
    a1 = _insert_account(conn, job_id, name="Cash",    balance=10000.0, number="1000")
    a2 = _insert_account(conn, job_id, name="Capital", balance=-10000.0, number="3000")
    return a1, a2


def _map_to_line(tmp_path, atbw_path, conn, job_id, account_ids,
                 fs="BalanceSheet", line_code="BS-01", line_name="Cash"):
    upsert_tax_line(conn, entity_type="1120S", financial_statement=fs,
                    line_code=line_code, line_name=line_name, sort_order=10)
    # need a second FS too so both-FS check passes — caller handles that
    tl = conn.execute(
        "SELECT tax_line_id FROM tax_lines WHERE line_code=? AND entity_type='1120S'",
        (line_code,),
    ).fetchone()
    map_accounts(conn, job_id=job_id, account_ids=account_ids,
                 tax_line_id=tl["tax_line_id"], mapped_by="preparer")


def _add_balanced_aje(conn, job_id, a1, a2):
    aje_id = uuid.uuid4().hex
    now = _now()
    conn.execute(
        """INSERT INTO journal_entries
               (aje_id, job_id, entry_type, entry_number, description,
                originated_by, originated_at, is_balanced, status)
           VALUES (?,?,'AJE','AJE-001','Test AJE','preparer',?,1,'Open')""",
        (aje_id, job_id, now),
    )
    for acct, amt in [(a1, 500.0), (a2, -500.0)]:
        conn.execute(
            """INSERT INTO journal_entry_lines
                   (line_id, aje_id, account_id, amount, sort_order)
               VALUES (?,?,?,?,0)""",
            (uuid.uuid4().hex, aje_id, acct, amt),
        )
    return aje_id


# ---------------------------------------------------------------------------
# Validation checks
# ---------------------------------------------------------------------------

def test_validation_fails_unbalanced_tb(atbw_path):
    job = get_job(atbw_path)
    with db_connection(atbw_path) as conn:
        _insert_account(conn, job["job_id"], name="Cash", balance=1000.0)
    with db_connection(atbw_path) as conn:
        results = run_all(conn, job["job_id"])
    labels = {r["label"]: r["status"] for r in results}
    assert labels["Trial balance is out of balance"] == "fail"


def test_validation_fails_unmapped_accounts(atbw_path):
    job = get_job(atbw_path)
    with db_connection(atbw_path) as conn:
        _insert_account(conn, job["job_id"], name="Cash", balance=0.0)
    with db_connection(atbw_path) as conn:
        results = run_all(conn, job["job_id"])
    fail_labels = [r["label"] for r in results if r["status"] == "fail"]
    assert any("unmapped" in l.lower() for l in fail_labels)


def test_validation_fails_unbalanced_aje(atbw_path):
    job = get_job(atbw_path)
    with db_connection(atbw_path) as conn:
        a1 = _insert_account(conn, job["job_id"], name="Cash", balance=0.0)
        aje_id = uuid.uuid4().hex
        now = _now()
        conn.execute(
            """INSERT INTO journal_entries
                   (aje_id, job_id, entry_type, entry_number, description,
                    originated_by, originated_at, is_balanced, status)
               VALUES (?,?,'AJE','AJE-001','Bad','preparer',?,0,'Shell')""",
            (aje_id, job["job_id"], now),
        )
        conn.execute(
            """INSERT INTO journal_entry_lines
                   (line_id, aje_id, account_id, amount, sort_order)
               VALUES (?,?,?,100.0,0)""",
            (uuid.uuid4().hex, aje_id, a1),
        )
    with db_connection(atbw_path) as conn:
        results = run_all(conn, job["job_id"])
    fail_labels = [r["label"] for r in results if r["status"] == "fail"]
    assert any("AJE" in l for l in fail_labels)


def test_validation_all_pass(tmp_path, atbw_path):
    db_path = tmp_path / "s.db"
    set_settings_path(db_path)
    ensure_settings_db()

    job = get_job(atbw_path)
    with db_connection(atbw_path) as conn:
        a1, a2 = _balanced_tb(conn, job["job_id"])
        # map a1 to BS, a2 to PL so both FS check passes
        upsert_tax_line(conn, entity_type="1120S", financial_statement="BalanceSheet",
                        line_code="BS-01", line_name="Cash", sort_order=10)
        upsert_tax_line(conn, entity_type="1120S", financial_statement="ProfitAndLoss",
                        line_code="PL-01", line_name="Revenue", sort_order=10)
        bs_tl = conn.execute(
            "SELECT tax_line_id FROM tax_lines WHERE line_code='BS-01'").fetchone()
        pl_tl = conn.execute(
            "SELECT tax_line_id FROM tax_lines WHERE line_code='PL-01'").fetchone()
        map_accounts(conn, job_id=job["job_id"], account_ids=[a1],
                     tax_line_id=bs_tl["tax_line_id"], mapped_by="preparer")
        map_accounts(conn, job_id=job["job_id"], account_ids=[a2],
                     tax_line_id=pl_tl["tax_line_id"], mapped_by="preparer")

    with db_connection(atbw_path) as conn:
        results = run_all(conn, job["job_id"])
    assert all_pass(results), [r for r in results if r["status"] == "fail"]


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def _setup_passing_binder(tmp_path, atbw_path):
    """Put the binder in a valid exportable state. Returns job dict."""
    db_path = tmp_path / "settings_exp.db"
    set_settings_path(db_path)
    ensure_settings_db()

    job = get_job(atbw_path)
    with db_connection(atbw_path) as conn:
        a1, a2 = _balanced_tb(conn, job["job_id"])
        upsert_tax_line(conn, entity_type="1120S", financial_statement="BalanceSheet",
                        line_code="BS-01", line_name="Cash", sort_order=10)
        upsert_tax_line(conn, entity_type="1120S", financial_statement="ProfitAndLoss",
                        line_code="PL-01", line_name="Revenue", sort_order=10)
        bs_tl = conn.execute(
            "SELECT tax_line_id FROM tax_lines WHERE line_code='BS-01'").fetchone()
        pl_tl = conn.execute(
            "SELECT tax_line_id FROM tax_lines WHERE line_code='PL-01'").fetchone()
        map_accounts(conn, job_id=job["job_id"], account_ids=[a1],
                     tax_line_id=bs_tl["tax_line_id"], mapped_by="preparer")
        map_accounts(conn, job_id=job["job_id"], account_ids=[a2],
                     tax_line_id=pl_tl["tax_line_id"], mapped_by="preparer")
        _add_balanced_aje(conn, job["job_id"], a1, a2)
    return get_job(atbw_path)


def test_export_creates_file(tmp_path, atbw_path):
    job = _setup_passing_binder(tmp_path, atbw_path)
    out = tmp_path / "test_review.atbr.xlsx"
    with db_connection(atbw_path) as conn:
        export_review_package(conn, job=job, output_path=out, performed_by="preparer")
    assert out.exists()
    assert out.stat().st_size > 0


def test_export_visible_tabs_present(tmp_path, atbw_path):
    job = _setup_passing_binder(tmp_path, atbw_path)
    out = tmp_path / "review_tabs.atbr.xlsx"
    with db_connection(atbw_path) as conn:
        export_review_package(conn, job=job, output_path=out, performed_by="preparer")
    wb = openpyxl.load_workbook(str(out))
    for expected in ("Cover", "Balance Sheet", "Income Statement", "Tax Grouping",
                     "Trial Balance", "Journal Entries", "Notes"):
        assert expected in wb.sheetnames, f"Missing tab: {expected}"


def test_export_no_activity_log_tab(tmp_path, atbw_path):
    """__activity_log should NOT appear in the export."""
    job = _setup_passing_binder(tmp_path, atbw_path)
    out = tmp_path / "review_no_log.atbr.xlsx"
    with db_connection(atbw_path) as conn:
        export_review_package(conn, job=job, output_path=out, performed_by="preparer")
    wb = openpyxl.load_workbook(str(out))
    assert "__activity_log" not in wb.sheetnames


def test_export_hidden_tabs_very_hidden(tmp_path, atbw_path):
    job = _setup_passing_binder(tmp_path, atbw_path)
    out = tmp_path / "review_hidden.atbr.xlsx"
    with db_connection(atbw_path) as conn:
        export_review_package(conn, job=job, output_path=out, performed_by="preparer")
    wb = openpyxl.load_workbook(str(out))
    for name in ("__manifest", "__data"):
        assert name in wb.sheetnames
        assert wb[name].sheet_state == "veryHidden"


def test_export_checksum_verifiable(tmp_path, atbw_path):
    """Checksum in __manifest should match SHA-256 of __data content only."""
    job = _setup_passing_binder(tmp_path, atbw_path)
    out = tmp_path / "review_checksum.atbr.xlsx"
    with db_connection(atbw_path) as conn:
        export_review_package(conn, job=job, output_path=out, performed_by="preparer")

    wb = openpyxl.load_workbook(str(out))
    manifest = {row[0].value: row[1].value
                for row in wb["__manifest"].iter_rows()}
    stored_checksum = manifest["checksum_sha256"]

    data_content = wb["__data"].cell(1, 1).value or ""
    expected = hashlib.sha256(data_content.encode("utf-8")).hexdigest()
    assert stored_checksum == expected


def test_export_cover_has_client_info(tmp_path, atbw_path):
    """Cover sheet should contain client name and preparer."""
    job = _setup_passing_binder(tmp_path, atbw_path)
    out = tmp_path / "review_cover.atbr.xlsx"
    with db_connection(atbw_path) as conn:
        export_review_package(conn, job=job, output_path=out, performed_by="preparer")
    wb = openpyxl.load_workbook(str(out))
    all_values = [str(cell.value or "") for row in wb["Cover"].iter_rows() for cell in row]
    assert any("ABC Company" in v for v in all_values)
    assert any("preparer" in v.lower() for v in all_values)


def test_export_writes_package_row(tmp_path, atbw_path):
    job = _setup_passing_binder(tmp_path, atbw_path)
    out = tmp_path / "review_pkg.atbr.xlsx"
    with db_connection(atbw_path) as conn:
        export_review_package(conn, job=job, output_path=out, performed_by="preparer")
    with db_connection(atbw_path) as conn:
        pkg = conn.execute(
            "SELECT * FROM packages WHERE job_id = ?", (job["job_id"],)
        ).fetchone()
    assert pkg is not None
    assert pkg["version_number"] == 1
    assert pkg["package_type"] == "review"


def test_export_version_increments(tmp_path, atbw_path):
    job = _setup_passing_binder(tmp_path, atbw_path)
    out1 = tmp_path / "v1.atbr.xlsx"
    out2 = tmp_path / "v2.atbr.xlsx"
    with db_connection(atbw_path) as conn:
        export_review_package(conn, job=job, output_path=out1, performed_by="preparer")
    with db_connection(atbw_path) as conn:
        export_review_package(conn, job=get_job(atbw_path), output_path=out2,
                              performed_by="preparer")
    with db_connection(atbw_path) as conn:
        rows = conn.execute(
            "SELECT version_number FROM packages WHERE job_id = ? ORDER BY version_number",
            (job["job_id"],),
        ).fetchall()
    assert [r[0] for r in rows] == [1, 2]


def test_export_updates_job_status(tmp_path, atbw_path):
    job = _setup_passing_binder(tmp_path, atbw_path)
    out = tmp_path / "review_status.atbr.xlsx"
    with db_connection(atbw_path) as conn:
        export_review_package(conn, job=job, output_path=out, performed_by="preparer")
    updated = get_job(atbw_path)
    assert updated["status"] == "Ready for Review"


def test_suggested_filename():
    job = {"tax_year": 2024, "client_name": "Acme Corp"}
    assert suggested_filename(job, 1) == "2024 Acme Corp Prep in Progress V01.atbr.xlsx"
    assert suggested_filename(job, 3) == "2024 Acme Corp Prep in Progress V03.atbr.xlsx"
