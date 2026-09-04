"""M8 tests: .atbr.xlsx as source of truth — round-trip and importer."""
import datetime
import hashlib
import json
import uuid
from pathlib import Path

import openpyxl
import pytest

from atbworkup.db.connection import db_connection
from atbworkup.db.settings import ensure_settings_db, set_settings_path
from atbworkup.exporter.review_package import save_workup, export_review_package
from atbworkup.importer.package import open_from_package
from atbworkup.models.job import create_workup, get_job
from atbworkup.models.mappings import map_accounts, upsert_tax_line
from atbworkup.utils.naming import suggested_filename, temp_atbw_path


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _now():
    return datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _insert_account(conn, job_id, *, name, balance=0.0, number=""):
    aid = uuid.uuid4().hex
    now = _now()
    conn.execute(
        """INSERT INTO accounts
               (account_id, job_id, account_number, account_name, account_type,
                pbc_balance, normal_balance, sort_order, is_mapped, created_at, updated_at)
           VALUES (?,?,?,?,?,?,?,?,0,?,?)""",
        (aid, job_id, number, name, "Asset", balance, "Debit", 0, now, now),
    )
    return aid


def _full_binder(tmp_path, settings_db):
    """
    Create a binder with two mapped accounts, one AJE, and one note.
    Returns (atbw_path, job).
    """
    set_settings_path(settings_db)
    ensure_settings_db()

    meta = {
        "client_name": "Round Trip Co",
        "entity_name": "Round Trip LLC",
        "tax_year": 2024,
        "entity_type": "1120S",
        "prepared_by": "Tester",
        "reviewer": None,
        "workpaper_folder": str(tmp_path),
        "accounting_system": None,
    }
    job_id = uuid.uuid4().hex
    atbw_path = tmp_path / "work.atbw"
    create_workup(atbw_path, meta, job_id=job_id)
    job = get_job(atbw_path)

    now = _now()
    with db_connection(atbw_path) as conn:
        a1 = _insert_account(conn, job_id, name="Cash",    balance=5000.0, number="1000")
        a2 = _insert_account(conn, job_id, name="Capital", balance=-5000.0, number="3000")

        upsert_tax_line(conn, entity_type="1120S", financial_statement="BalanceSheet",
                        line_code="BS-01", line_name="Cash", sort_order=10)
        upsert_tax_line(conn, entity_type="1120S", financial_statement="BalanceSheet",
                        line_code="BS-50", line_name="Equity", sort_order=50)

        tl_cash = conn.execute(
            "SELECT tax_line_id FROM tax_lines WHERE line_code='BS-01' AND entity_type='1120S'"
        ).fetchone()["tax_line_id"]
        tl_eq = conn.execute(
            "SELECT tax_line_id FROM tax_lines WHERE line_code='BS-50' AND entity_type='1120S'"
        ).fetchone()["tax_line_id"]

        map_accounts(conn, job_id=job_id, account_ids=[a1], tax_line_id=tl_cash, mapped_by="preparer")
        map_accounts(conn, job_id=job_id, account_ids=[a2], tax_line_id=tl_eq, mapped_by="preparer")

        aje_id = uuid.uuid4().hex
        conn.execute(
            """INSERT INTO journal_entries
                   (aje_id, job_id, entry_type, entry_number, description,
                    originated_by, originated_at, is_balanced, status)
               VALUES (?,?,'AJE','AJE-001','Adjustment','preparer',?,1,'Open')""",
            (aje_id, job_id, now),
        )
        line1 = uuid.uuid4().hex
        line2 = uuid.uuid4().hex
        conn.execute(
            "INSERT INTO journal_entry_lines (line_id, aje_id, account_id, amount, sort_order) VALUES (?,?,?,?,0)",
            (line1, aje_id, a1, 100.0),
        )
        conn.execute(
            "INSERT INTO journal_entry_lines (line_id, aje_id, account_id, amount, sort_order) VALUES (?,?,?,?,1)",
            (line2, aje_id, a2, -100.0),
        )

        note_id = uuid.uuid4().hex
        conn.execute(
            """INSERT INTO notes
                   (note_id, job_id, note_type, linked_to_type, linked_to_id,
                    body, created_by, created_at, status)
               VALUES (?,?,'preparer','account',?,?,?,?,?)""",
            (note_id, job_id, a1, "Check this balance.", "Tester", now, "Open"),
        )

    return atbw_path, get_job(atbw_path)


# ---------------------------------------------------------------------------
# suggested_filename
# ---------------------------------------------------------------------------

def test_suggested_filename_extension():
    name = suggested_filename(2025, "ABC Co")
    assert name.endswith(".atbr.xlsx")


def test_temp_atbw_path_uses_job_id():
    job_id = uuid.uuid4().hex
    p = temp_atbw_path(job_id)
    assert p.name == f"{job_id}.atbw"
    assert "ATBWorkup" in str(p)


# ---------------------------------------------------------------------------
# save_workup — writes a valid xlsx without incrementing version
# ---------------------------------------------------------------------------

def test_save_workup_creates_file(tmp_path, meta):
    atbw = tmp_path / "work.atbw"
    create_workup(atbw, meta)
    job = get_job(atbw)
    out = tmp_path / "output.atbr.xlsx"

    with db_connection(atbw) as conn:
        save_workup(conn, job=job, output_path=out, performed_by="Tester")

    assert out.exists()


def test_save_workup_has_required_tabs(tmp_path, meta):
    atbw = tmp_path / "work.atbw"
    create_workup(atbw, meta)
    job = get_job(atbw)
    out = tmp_path / "output.atbr.xlsx"

    with db_connection(atbw) as conn:
        save_workup(conn, job=job, output_path=out, performed_by="Tester")

    wb = openpyxl.load_workbook(str(out))
    visible = [ws.title for ws in wb.worksheets if not ws.title.startswith("__")]
    assert "Cover" in visible
    assert "Balance Sheet" in visible
    assert "Income Statement" in visible


def test_save_workup_does_not_write_packages_row(tmp_path, meta):
    atbw = tmp_path / "work.atbw"
    create_workup(atbw, meta)
    job = get_job(atbw)
    out = tmp_path / "output.atbr.xlsx"

    with db_connection(atbw) as conn:
        save_workup(conn, job=job, output_path=out, performed_by="Tester")
        count = conn.execute("SELECT COUNT(*) FROM packages").fetchone()[0]

    assert count == 0


def test_save_workup_version_stays_at_1_for_new_binder(tmp_path, meta):
    atbw = tmp_path / "work.atbw"
    create_workup(atbw, meta)
    job = get_job(atbw)
    out = tmp_path / "output.atbr.xlsx"

    with db_connection(atbw) as conn:
        save_workup(conn, job=job, output_path=out, performed_by="Tester")

    wb = openpyxl.load_workbook(str(out), data_only=True)
    ws = wb["__manifest"]
    manifest = {row[0]: row[1] for row in ws.iter_rows(values_only=True) if row[0]}
    assert str(manifest["version_number"]) == "1"


# ---------------------------------------------------------------------------
# __data snapshot — schema_version 2.0 and full content
# ---------------------------------------------------------------------------

def test_data_tab_schema_version_2(tmp_path):
    settings_db = tmp_path / "settings.db"
    atbw_path, job = _full_binder(tmp_path, settings_db)
    out = tmp_path / "out.atbr.xlsx"

    with db_connection(atbw_path) as conn:
        save_workup(conn, job=job, output_path=out, performed_by="Tester")

    wb = openpyxl.load_workbook(str(out), data_only=True)
    data = json.loads(wb["__data"].cell(1, 1).value)
    assert data["schema_version"] == "2.0"


def test_data_tab_contains_full_snapshot(tmp_path):
    settings_db = tmp_path / "settings.db"
    atbw_path, job = _full_binder(tmp_path, settings_db)
    out = tmp_path / "out.atbr.xlsx"

    with db_connection(atbw_path) as conn:
        save_workup(conn, job=job, output_path=out, performed_by="Tester")

    wb = openpyxl.load_workbook(str(out), data_only=True)
    data = json.loads(wb["__data"].cell(1, 1).value)

    assert len(data["accounts"]) == 2
    assert len(data["tax_lines"]) >= 2
    assert len(data["mappings"]) == 2
    assert len(data["entries"]) == 1
    assert len(data["entry_lines"]) == 2
    assert len(data["notes"]) == 1


def test_data_tab_checksum_verifiable(tmp_path):
    settings_db = tmp_path / "settings.db"
    atbw_path, job = _full_binder(tmp_path, settings_db)
    out = tmp_path / "out.atbr.xlsx"

    with db_connection(atbw_path) as conn:
        save_workup(conn, job=job, output_path=out, performed_by="Tester")

    wb = openpyxl.load_workbook(str(out), data_only=True)
    data_content = wb["__data"].cell(1, 1).value
    actual = hashlib.sha256(data_content.encode("utf-8")).hexdigest()
    manifest = {row[0]: row[1] for row in wb["__manifest"].iter_rows(values_only=True) if row[0]}
    assert manifest["checksum_sha256"] == actual


# ---------------------------------------------------------------------------
# open_from_package — full round-trip
# ---------------------------------------------------------------------------

def test_open_from_package_returns_job(tmp_path):
    settings_db = tmp_path / "settings.db"
    atbw_path, job = _full_binder(tmp_path, settings_db)
    out = tmp_path / "package.atbr.xlsx"

    with db_connection(atbw_path) as conn:
        save_workup(conn, job=job, output_path=out, performed_by="Tester")

    temp_path, reimported_job = open_from_package(out, performed_by="Reviewer")
    assert reimported_job["client_name"] == "Round Trip Co"
    assert reimported_job["tax_year"] == 2024
    assert reimported_job["job_id"] == job["job_id"]

    temp_path.unlink(missing_ok=True)


def test_open_from_package_hydrates_accounts(tmp_path):
    settings_db = tmp_path / "settings.db"
    atbw_path, job = _full_binder(tmp_path, settings_db)
    out = tmp_path / "package.atbr.xlsx"

    with db_connection(atbw_path) as conn:
        save_workup(conn, job=job, output_path=out, performed_by="Tester")

    temp_path, _ = open_from_package(out, performed_by="Reviewer")
    with db_connection(temp_path) as conn:
        count = conn.execute(
            "SELECT COUNT(*) FROM accounts WHERE job_id = ?", (job["job_id"],)
        ).fetchone()[0]
    assert count == 2
    temp_path.unlink(missing_ok=True)


def test_open_from_package_hydrates_entries_and_lines(tmp_path):
    settings_db = tmp_path / "settings.db"
    atbw_path, job = _full_binder(tmp_path, settings_db)
    out = tmp_path / "package.atbr.xlsx"

    with db_connection(atbw_path) as conn:
        save_workup(conn, job=job, output_path=out, performed_by="Tester")

    temp_path, _ = open_from_package(out, performed_by="Reviewer")
    with db_connection(temp_path) as conn:
        entries = conn.execute(
            "SELECT COUNT(*) FROM journal_entries WHERE job_id = ?", (job["job_id"],)
        ).fetchone()[0]
        lines = conn.execute(
            """SELECT COUNT(*) FROM journal_entry_lines jel
               JOIN journal_entries je ON je.aje_id = jel.aje_id
               WHERE je.job_id = ?""",
            (job["job_id"],),
        ).fetchone()[0]
    assert entries == 1
    assert lines == 2
    temp_path.unlink(missing_ok=True)


def test_open_from_package_hydrates_notes(tmp_path):
    settings_db = tmp_path / "settings.db"
    atbw_path, job = _full_binder(tmp_path, settings_db)
    out = tmp_path / "package.atbr.xlsx"

    with db_connection(atbw_path) as conn:
        save_workup(conn, job=job, output_path=out, performed_by="Tester")

    temp_path, _ = open_from_package(out, performed_by="Reviewer")
    with db_connection(temp_path) as conn:
        notes = conn.execute(
            "SELECT COUNT(*) FROM notes WHERE job_id = ?", (job["job_id"],)
        ).fetchone()[0]
    assert notes == 1
    temp_path.unlink(missing_ok=True)


def test_open_from_package_hydrates_mappings(tmp_path):
    settings_db = tmp_path / "settings.db"
    atbw_path, job = _full_binder(tmp_path, settings_db)
    out = tmp_path / "package.atbr.xlsx"

    with db_connection(atbw_path) as conn:
        save_workup(conn, job=job, output_path=out, performed_by="Tester")

    temp_path, _ = open_from_package(out, performed_by="Reviewer")
    with db_connection(temp_path) as conn:
        mappings = conn.execute(
            "SELECT COUNT(*) FROM mappings WHERE job_id = ?", (job["job_id"],)
        ).fetchone()[0]
    assert mappings == 2
    temp_path.unlink(missing_ok=True)


def test_open_from_package_rejects_tampered_file(tmp_path):
    settings_db = tmp_path / "settings.db"
    atbw_path, job = _full_binder(tmp_path, settings_db)
    out = tmp_path / "package.atbr.xlsx"

    with db_connection(atbw_path) as conn:
        save_workup(conn, job=job, output_path=out, performed_by="Tester")

    # Tamper: overwrite __data content directly via openpyxl
    import openpyxl as xl
    wb = xl.load_workbook(str(out))
    wb["__data"].sheet_state = "visible"
    wb["__data"].cell(1, 1).value = '{"tampered": true}'
    wb.save(str(out))

    with pytest.raises(ValueError, match="checksum"):
        open_from_package(out, performed_by="Attacker")


def test_open_from_package_rejects_missing_manifest(tmp_path):
    out = tmp_path / "not_a_package.atbr.xlsx"
    import openpyxl as xl
    wb = xl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(str(out))

    with pytest.raises(ValueError, match="__manifest"):
        open_from_package(out, performed_by="Anyone")


def test_create_workup_preserves_job_id(tmp_path, meta):
    custom_id = uuid.uuid4().hex
    atbw = tmp_path / "work.atbw"
    returned_id = create_workup(atbw, meta, job_id=custom_id)
    assert returned_id == custom_id
    job = get_job(atbw)
    assert job["job_id"] == custom_id
