"""
M9 tests: consolidation math + export/UI parity + save robustness.

These lock in the exact bugs found and fixed in the July 6 consolidation
session so they can't silently regress:
  - net income must be -sum(raw), not a blind sum of section displays
    (blind sum double-adds expenses instead of subtracting them)
  - Schedule K sections mix income- and deduction-type accounts in one
    section, so they need the same raw-based treatment, keyed off the
    stored tax_line.category, not a section-name guess
  - account-level elimination/CTE lines must fold into net income with the
    correct sign based on each account's own normal_balance
  - the Excel export and the interactive window must produce identical
    numbers, since compute_combined() is the only place either is allowed
    to do this math
  - a consolidated job's export must use the combined-subsidiary tab
    writers, not the regular (always-empty-for-consolidated) ones
  - saves must be atomic: an interrupted write must never corrupt the
    target file
"""
import uuid
import datetime
import zipfile

import openpyxl

from atbworkup.db.connection import db_connection
from atbworkup.models.job import create_workup, get_job
from atbworkup.models.accounts import create_account
from atbworkup.models.mappings import upsert_tax_line, map_accounts
from atbworkup.models import consolidation_calc
from atbworkup.exporter.review_package import save_workup, _atomic_save
from atbworkup.data.tax_line_categories import CATEGORY_SCHEDULE_K


def _now():
    return datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _make_subsidiary(tmp_path, name, accounts, entity_type="1120S"):
    """
    Build a subsidiary .atbw, map the given accounts, export it to a
    .atbr.xlsx, and return that path.

    accounts: list of (number, account_name, account_type, normal_balance,
                        pbc_balance, financial_statement, section, line_code,
                        line_name, category_or_None)
    """
    job_id = uuid.uuid4().hex
    atbw_path = tmp_path / f"{name}.atbw"
    meta = {
        "client_name": name, "entity_name": name, "tax_year": 2025,
        "entity_type": entity_type, "prepared_by": "Test Preparer",
        "workpaper_folder": str(tmp_path), "accounting_system": "QuickBooks",
    }
    create_workup(atbw_path, meta, job_id=job_id)

    with db_connection(atbw_path) as conn:
        for (num, aname, atype, nb, bal, fs, section, code, lname, category) in accounts:
            tl_id = upsert_tax_line(
                conn, entity_type=entity_type, financial_statement=fs,
                section=section, section_sort_order=10, line_code=code,
                line_name=lname, sort_order=10, category=category or "",
            )
            aid = create_account(conn, job_id=job_id, account_number=num,
                                 account_name=aname, account_type=atype,
                                 normal_balance=nb, pbc_balance=bal)
            map_accounts(conn, job_id=job_id, account_ids=[aid],
                        tax_line_id=tl_id, mapped_by="preparer")

    job = get_job(atbw_path)
    out_path = tmp_path / f"{name}.atbr.xlsx"
    with db_connection(atbw_path) as conn:
        save_workup(conn, job=job, output_path=out_path, performed_by="preparer")
    return out_path


def _make_consolidated_job(tmp_path, members):
    """members: list of (label, code, atbr_xlsx_path). Returns (conn ctx path, job)."""
    job_id = uuid.uuid4().hex
    atbw_path = tmp_path / "consolidated.atbw"
    meta = {
        "client_name": "Test Group (Consolidated)", "entity_name": "Test Group (Consolidated)",
        "tax_year": 2025, "entity_type": "Consolidated", "prepared_by": "Test Preparer",
        "workpaper_folder": str(tmp_path), "accounting_system": "QuickBooks",
    }
    create_workup(atbw_path, meta, job_id=job_id)
    with db_connection(atbw_path) as conn:
        for i, (label, code, path) in enumerate(members):
            conn.execute(
                "INSERT INTO consolidation_members "
                "(member_id, job_id, member_name, member_code, file_path, "
                " member_type, sort_order, created_at) VALUES (?,?,?,?,?,?,?,?)",
                (uuid.uuid4().hex, job_id, label, code, str(path), "subsidiary", i, _now()),
            )
    return atbw_path, get_job(atbw_path)


# ---------------------------------------------------------------------------
# Net income must be raw-sign-based, not a blind section-display sum
# ---------------------------------------------------------------------------

def test_net_income_is_not_blind_sum_of_sections(tmp_path):
    """Revenue 100, COGS 30, Deductions 20 -> net income 50, not 150."""
    sub = _make_subsidiary(tmp_path, "Sub1", [
        ("4000", "Sales", "Revenue", "Credit", -100.0,
         "ProfitAndLoss", "Revenue", "PL-01", "Sales", None),
        ("5000", "COGS", "Expense", "Debit", 30.0,
         "ProfitAndLoss", "Cost of Goods Sold", "PL-02", "COGS", None),
        ("6000", "Rent", "Expense", "Debit", 20.0,
         "ProfitAndLoss", "Deductions", "PL-03", "Rent", None),
    ])
    atbw_path, job = _make_consolidated_job(tmp_path, [("Sub1", "S1", sub)])
    with db_connection(atbw_path) as conn:
        calc = consolidation_calc.compute_combined(conn, job)
    assert calc["combined_net_income"] == 50.0


def test_category_auto_classification(tmp_path):
    job_id = uuid.uuid4().hex
    atbw_path = tmp_path / "probe.atbw"
    meta = {
        "client_name": "Probe", "entity_name": "Probe", "tax_year": 2025,
        "entity_type": "1120S", "prepared_by": "Test", "workpaper_folder": str(tmp_path),
        "accounting_system": "QuickBooks",
    }
    create_workup(atbw_path, meta, job_id=job_id)
    with db_connection(atbw_path) as conn:
        upsert_tax_line(conn, entity_type="1120S", financial_statement="ProfitAndLoss",
                        section="Revenue", line_code="PL-01", line_name="Sales", sort_order=10)
        upsert_tax_line(conn, entity_type="1120S", financial_statement="ProfitAndLoss",
                        section="Cost of Goods Sold", line_code="PL-02", line_name="COGS", sort_order=20)
        rows = {r["line_code"]: r["category"] for r in conn.execute("SELECT line_code, category FROM tax_lines")}
    assert rows["PL-01"] == "revenue"
    assert rows["PL-02"] == "cogs"


# ---------------------------------------------------------------------------
# Schedule K: mixed income-type and deduction-type accounts in one section
# ---------------------------------------------------------------------------

def test_schedule_k_mixed_sign_items_net_correctly(tmp_path):
    """Interest income +500 (credit) and charitable -200 (debit) -> net +300,
    and it must be broken out separately from Ordinary Business Income."""
    sub = _make_subsidiary(tmp_path, "KSub", [
        ("9100", "Interest Income", "Revenue", "Credit", -500.0,
         "ProfitAndLoss", "Schedule K — Pass-Through", "K-4", "Interest Income",
         CATEGORY_SCHEDULE_K),
        ("9200", "Charitable Contributions", "Expense", "Debit", 200.0,
         "ProfitAndLoss", "Schedule K — Pass-Through", "K-12a", "Charitable Contributions",
         CATEGORY_SCHEDULE_K),
    ])
    atbw_path, job = _make_consolidated_job(tmp_path, [("KSub", "KS", sub)])
    with db_connection(atbw_path) as conn:
        calc = consolidation_calc.compute_combined(conn, job)
    assert calc["has_sch_k"] is True
    assert calc["sch_k_total_ni"] == 300.0
    assert calc["combined_net_income"] == 300.0


# ---------------------------------------------------------------------------
# Account-level eliminations must fold in with the correct sign
# ---------------------------------------------------------------------------

def test_account_level_elimination_debit_to_expense_reduces_net_income(tmp_path):
    """A $1,000 DEBIT elimination against an EXPENSE account must SUBTRACT
    1,000 from net income, not add it — this was the exact sign bug found
    and fixed this session."""
    from atbworkup.models import consolidation_entries as ce_model
    from atbworkup.models import consolidation_read

    sub = _make_subsidiary(tmp_path, "ExpSub", [
        ("6100", "Consulting Expense", "Expense", "Debit", 500.0,
         "ProfitAndLoss", "Deductions", "PL-01", "Consulting", None),
    ])
    atbw_path, job = _make_consolidated_job(tmp_path, [("ExpSub", "ES", sub)])

    idx = consolidation_read.read_member_account_index(sub)
    expense_account_id = next(iter(idx.keys()))

    with db_connection(atbw_path) as conn:
        member_id = conn.execute(
            "SELECT member_id FROM consolidation_members WHERE job_id=?", (job["job_id"],)
        ).fetchone()["member_id"]
        base_ni = consolidation_calc.compute_combined(conn, job)["combined_net_income"]

        entry = ce_model.create_entry(conn, job_id=job["job_id"], workpaper="elim",
                                      description="test", originated_by="preparer")
        ce_model.save_lines(conn, entry["entry_id"], [
            {"member_id": member_id, "account_id": expense_account_id,
             "amount": 1000.0, "memo": ""},
        ])
        after_ni = consolidation_calc.compute_combined(conn, job)["combined_net_income"]

    assert base_ni == -500.0     # $500 expense reduces net income
    assert after_ni == base_ni - 1000.0   # additional $1,000 debit to the SAME expense


# ---------------------------------------------------------------------------
# K-1 allocation query must actually match real Schedule K tax lines.
# Found in passing: it filtered on financial_statement = 'ScheduleK', a
# value nothing is ever seeded with (real K-lines use financial_statement =
# 'ProfitAndLoss' + category = 'schedule_k') — the table always showed
# "no accounts mapped" even when K-1 accounts existed.
# ---------------------------------------------------------------------------

def test_k1_allocation_query_matches_real_schedule_k_lines(tmp_path):
    job_id = uuid.uuid4().hex
    atbw_path = tmp_path / "k1.atbw"
    meta = {
        "client_name": "K1 Co", "entity_name": "K1 Co", "tax_year": 2025,
        "entity_type": "1120S", "prepared_by": "Test", "workpaper_folder": str(tmp_path),
        "accounting_system": "QuickBooks",
    }
    create_workup(atbw_path, meta, job_id=job_id)
    with db_connection(atbw_path) as conn:
        tl_id = upsert_tax_line(conn, entity_type="1120S", financial_statement="ProfitAndLoss",
                                section="Schedule K — Pass-Through", line_code="K-4",
                                line_name="Interest Income", sort_order=10)
        aid = create_account(conn, job_id=job_id, account_number="9100",
                             account_name="Interest Income", account_type="Revenue",
                             normal_balance="Credit", pbc_balance=-500.0)
        map_accounts(conn, job_id=job_id, account_ids=[aid], tax_line_id=tl_id,
                    mapped_by="preparer")

        rows = conn.execute(
            """
            SELECT tl.line_code, tl.line_name, a.pbc_balance
            FROM accounts a
            JOIN mappings m ON m.account_id = a.account_id AND m.job_id = a.job_id
            JOIN tax_lines tl ON tl.tax_line_id = m.tax_line_id
                AND tl.category = ?
            WHERE a.job_id = ?
            """,
            (CATEGORY_SCHEDULE_K, job_id),
        ).fetchall()
    assert len(rows) == 1
    assert rows[0]["line_code"] == "K-4"


# ---------------------------------------------------------------------------
# Export must match the interactive calculation exactly (no duplicate math)
# ---------------------------------------------------------------------------

def test_consolidated_export_matches_compute_combined(tmp_path):
    sub = _make_subsidiary(tmp_path, "ParitySub", [
        ("4000", "Sales", "Revenue", "Credit", -1000.0,
         "ProfitAndLoss", "Revenue", "PL-01", "Sales", None),
        ("5000", "COGS", "Expense", "Debit", 300.0,
         "ProfitAndLoss", "Cost of Goods Sold", "PL-02", "COGS", None),
    ])
    atbw_path, job = _make_consolidated_job(tmp_path, [("ParitySub", "PS", sub)])

    with db_connection(atbw_path) as conn:
        calc = consolidation_calc.compute_combined(conn, job)
        out = tmp_path / "parity.atbr.xlsx"
        save_workup(conn, job=job, output_path=out, performed_by="preparer")

    wb = openpyxl.load_workbook(str(out))
    ws = wb["Combined Income Statement"]
    values = {row[0]: row[1] for row in ws.iter_rows(values_only=True) if row[0]}
    assert values["Net Income / (Loss)"] == calc["combined_net_income"] == 700.0
    assert values["Gross Profit"] == 700.0


def test_consolidated_job_gets_combined_tabs_not_empty_regular_tabs(tmp_path):
    """Regression guard for 'reports read as empty' — a Consolidated job's
    export must use the combined-subsidiary writers."""
    sub = _make_subsidiary(tmp_path, "TabSub", [
        ("4000", "Sales", "Revenue", "Credit", -100.0,
         "ProfitAndLoss", "Revenue", "PL-01", "Sales", None),
    ])
    atbw_path, job = _make_consolidated_job(tmp_path, [("TabSub", "TS", sub)])
    out = tmp_path / "tabs.atbr.xlsx"
    with db_connection(atbw_path) as conn:
        save_workup(conn, job=job, output_path=out, performed_by="preparer")

    wb = openpyxl.load_workbook(str(out))
    assert "Combined Balance Sheet" in wb.sheetnames
    assert "Combined Income Statement" in wb.sheetnames
    assert "Combined Tax Grouping" in wb.sheetnames
    assert "Balance Sheet" not in wb.sheetnames   # regular tab must NOT appear

    ws = wb["Combined Tax Grouping"]
    rows = [r for r in ws.iter_rows(values_only=True) if r[0]]
    assert len(rows) > 1, "Combined Tax Grouping tab is empty"


def test_regular_job_still_gets_regular_tabs(tmp_path):
    """Non-consolidated jobs must be completely unaffected."""
    job_id = uuid.uuid4().hex
    atbw_path = tmp_path / "regular.atbw"
    meta = {
        "client_name": "Regular Co", "entity_name": "Regular Co", "tax_year": 2025,
        "entity_type": "1120S", "prepared_by": "Test", "workpaper_folder": str(tmp_path),
        "accounting_system": "QuickBooks",
    }
    create_workup(atbw_path, meta, job_id=job_id)
    job = get_job(atbw_path)
    out = tmp_path / "regular.atbr.xlsx"
    with db_connection(atbw_path) as conn:
        save_workup(conn, job=job, output_path=out, performed_by="preparer")

    wb = openpyxl.load_workbook(str(out))
    assert "Balance Sheet" in wb.sheetnames
    assert "Combined Balance Sheet" not in wb.sheetnames


# ---------------------------------------------------------------------------
# Save robustness
# ---------------------------------------------------------------------------

def test_atomic_save_produces_valid_zip(tmp_path):
    import openpyxl as _oxl
    wb = _oxl.Workbook()
    out = tmp_path / "atomic.xlsx"
    _atomic_save(wb, out)
    with zipfile.ZipFile(out) as z:
        assert z.testzip() is None


def test_atomic_save_leaves_no_temp_files(tmp_path):
    import openpyxl as _oxl
    wb = _oxl.Workbook()
    out = tmp_path / "atomic2.xlsx"
    _atomic_save(wb, out)
    leftovers = list(tmp_path.glob(".*atomic2*.tmp"))
    assert leftovers == []


def test_atomic_save_preserves_original_on_failure(tmp_path, monkeypatch):
    """If the write itself fails, the original file must be untouched."""
    import openpyxl as _oxl

    out = tmp_path / "existing.xlsx"
    good_wb = _oxl.Workbook()
    _atomic_save(good_wb, out)
    original_bytes = out.read_bytes()

    class _BoomWorkbook:
        def save(self, path):
            raise RuntimeError("simulated interrupted write")

    try:
        _atomic_save(_BoomWorkbook(), out)
    except RuntimeError:
        pass

    assert out.read_bytes() == original_bytes
